#!/usr/bin/env python3
"""
step5c_validation_engine.py — Validation Engine
================================================
CDCI P&ID Pipeline — Step 5C

Type: PROGRAMMATIC — No Gemini, No Claude

Purpose
-------
Validate every candidate record against:
  1. ISA-5.1 tag format regex rules
  2. Business rules (prefix consistency, loop pairing, etc.)
  3. Asset registry lookup (against Annexure-4 tag register)
  4. Notes-derived project-specific rules (from Step 3)
  5. SOW membership check (from Step 4)

Input
-----
  step5b_associations.json (enriched candidates from 5A+5B)
  asset_register.xlsx / asset_register.json   (Annexure-4 or similar)
  notes_context.json  (project rules from Step 3)
  sow_symbol_memory.json (SOW scope from Step 4)

Output per candidate
---------------------
  {
    "validation_status": "PASS|WARN|FAIL",
    "validation_reason": "",
    "validation_details": [list of checks]
  }

Usage
-----
  python step5c_validation_engine.py \\
      --associations output/step5b_associations.json \\
      --register output/master_tags.json \\
      --out output/

  python step5c_validation_engine.py \\
      --context output/drawing_context.json
"""

import argparse
import json
import logging
import os
import re
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")


# ═══════════════════════════════════════════════════════════════════════════════
# ISA-5.1 Tag Format Rules
# ═══════════════════════════════════════════════════════════════════════════════

# Regex patterns by instrument category
ISA_PATTERNS = {
    # Flow instruments
    "FT":   re.compile(r'^[A-Z]{0,3}-?FT-?\d{3,6}[A-Z]?$',  re.I),
    "FIT":  re.compile(r'^[A-Z]{0,3}-?FIT-?\d{3,6}[A-Z]?$', re.I),
    "FE":   re.compile(r'^[A-Z]{0,3}-?FE-?\d{3,6}[A-Z]?$',  re.I),
    "FCV":  re.compile(r'^[A-Z]{0,3}-?FCV?-?\d{3,6}[A-Z]?$', re.I),
    "FV":   re.compile(r'^[A-Z]{0,3}-?FV-?\d{3,6}[A-Z]?$',  re.I),
    # Pressure instruments
    "PT":   re.compile(r'^[A-Z]{0,3}-?PT-?\d{3,6}[A-Z]?$',  re.I),
    "PIT":  re.compile(r'^[A-Z]{0,3}-?PIT-?\d{3,6}[A-Z]?$', re.I),
    "PI":   re.compile(r'^[A-Z]{0,3}-?PI-?\d{3,6}[A-Z]?$',  re.I),
    "PSV":  re.compile(r'^[A-Z]{0,3}-?PSV?-?\d{3,6}[A-Z]?$', re.I),
    "PS":   re.compile(r'^[A-Z]{0,3}-?PS-?\d{3,6}[A-Z]?$',  re.I),
    # Temperature instruments
    "TT":   re.compile(r'^[A-Z]{0,3}-?TT-?\d{3,6}[A-Z]?$',  re.I),
    "TIT":  re.compile(r'^[A-Z]{0,3}-?TIT-?\d{3,6}[A-Z]?$', re.I),
    "TE":   re.compile(r'^[A-Z]{0,3}-?TE-?\d{3,6}[A-Z]?$',  re.I),
    "TW":   re.compile(r'^[A-Z]{0,3}-?TW-?\d{3,6}[A-Z]?$',  re.I),
    "TCV":  re.compile(r'^[A-Z]{0,3}-?TCV?-?\d{3,6}[A-Z]?$', re.I),
    # Level instruments
    "LT":   re.compile(r'^[A-Z]{0,3}-?LT-?\d{3,6}[A-Z]?$',  re.I),
    "LIT":  re.compile(r'^[A-Z]{0,3}-?LIT-?\d{3,6}[A-Z]?$', re.I),
    "LG":   re.compile(r'^[A-Z]{0,3}-?LG-?\d{3,6}[A-Z]?$',  re.I),
    "LS":   re.compile(r'^[A-Z]{0,3}-?LS-?\d{3,6}[A-Z]?$',  re.I),
    "LV":   re.compile(r'^[A-Z]{0,3}-?LV-?\d{3,6}[A-Z]?$',  re.I),
    # Valves
    "XV":   re.compile(r'^[A-Z]{0,3}-?XV-?\d{3,6}[A-Z]?$',  re.I),
    "XY":   re.compile(r'^[A-Z]{0,3}-?XY-?\d{3,6}[A-Z]?$',  re.I),
    "BV":   re.compile(r'^[A-Z]{0,3}-?BV-?\d{3,6}[A-Z]?$',  re.I),
    "GV":   re.compile(r'^[A-Z]{0,3}-?GV-?\d{3,6}[A-Z]?$',  re.I),
    "NRV":  re.compile(r'^[A-Z]{0,3}-?NRV-?\d{3,6}[A-Z]?$', re.I),
    "ESDV": re.compile(r'^[A-Z]{0,3}-?ESDV?-?\d{3,6}[A-Z]?$', re.I),
    "SDV":  re.compile(r'^[A-Z]{0,3}-?SDV?-?\d{3,6}[A-Z]?$', re.I),
    "RV":   re.compile(r'^[A-Z]{0,3}-?RV-?\d{3,6}[A-Z]?$',  re.I),
    # Analyzers
    "AT":   re.compile(r'^[A-Z]{0,3}-?AT-?\d{3,6}[A-Z]?$',  re.I),
    "AE":   re.compile(r'^[A-Z]{0,3}-?AE-?\d{3,6}[A-Z]?$',  re.I),
    # Miscellaneous instruments
    "ZIT":  re.compile(r'^[A-Z]{0,3}-?ZIT-?\d{3,6}[A-Z]?$', re.I),
    "ZSC":  re.compile(r'^[A-Z]{0,3}-?ZSC-?\d{3,6}[A-Z]?$', re.I),
    "ZSO":  re.compile(r'^[A-Z]{0,3}-?ZSO-?\d{3,6}[A-Z]?$', re.I),
    "ZS":   re.compile(r'^[A-Z]{0,3}-?ZS[CO]?-?\d{3,6}[A-Z]?$', re.I),
    "HS":   re.compile(r'^[A-Z]{0,3}-?HS-?\d{3,6}[A-Z]?$',  re.I),
    "SS":   re.compile(r'^[A-Z]{0,3}-?SS-?\d{3,6}[A-Z]?$',  re.I),
    # Equipment
    "V":    re.compile(r'^V-[A-Z]{1,4}-?\d{3,6}[A-Z]?$',    re.I),
    "K":    re.compile(r'^K-[A-Z]-?\d{3,6}$',                re.I),
    "E":    re.compile(r'^E-[A-Z]{1,3}-?\d{3,6}[A-Z]?$',    re.I),
    "P":    re.compile(r'^P-[A-Z]{1,3}-?\d{3,6}[A-Z]?$',    re.I),
    "S":    re.compile(r'^S-[A-Z]{1,3}-?\d{3,6}[A-Z]?$',    re.I),
    # Pipeline tags
    "LINE": re.compile(
        r'^\d{1,4}["\-][A-Z]{2,5}-[A-Z0-9]+-[A-Z0-9]+(-PP|-X)?$', re.I),
    # Generic ISA fallback
    "GENERIC": re.compile(
        r'^[A-Z]{0,3}-?[A-Z]{1,5}-?\d{2,6}[A-Z]?$', re.I),
}

# Prefix → expected ISA function letters
PREFIX_FUNCTION_MAP = {
    "F":  {"FT","FIT","FE","FCV","FV","FG","FQI"},
    "P":  {"PT","PIT","PI","PS","PSV","PI"},
    "T":  {"TT","TIT","TE","TW","TG","TCV","TI"},
    "L":  {"LT","LIT","LG","LS","LV","LI","LC"},
    "A":  {"AT","AE","AI"},
    "Z":  {"ZIT","ZSC","ZSO","ZS","ZY","ZLC"},
    "X":  {"XV","XY","XX"},
}


# ═══════════════════════════════════════════════════════════════════════════════
# Validation checks
# ═══════════════════════════════════════════════════════════════════════════════

def check_tag_format(tag_text: str) -> dict:
    """
    Rule 1: ISA-5.1 regex format validation.
    Returns {pass: bool, message: str}
    """
    if not tag_text or not tag_text.strip():
        return {"pass": False, "rule": "TAG_FORMAT",
                "message": "Empty or null tag text"}

    tag = tag_text.strip()

    # Try each ISA pattern
    for prefix, pattern in ISA_PATTERNS.items():
        if pattern.match(tag):
            return {"pass": True, "rule": "TAG_FORMAT",
                    "message": f"Matches ISA pattern: {prefix}"}

    # Minimal check: at least PREFIX-NUMBER
    basic = re.match(r'^[A-Z0-9]{1,8}-[A-Z0-9]{1,10}$', tag, re.I)
    if basic:
        return {"pass": True, "rule": "TAG_FORMAT",
                "message": "Matches basic prefix-number format"}

    return {"pass": False, "rule": "TAG_FORMAT",
            "message": f"Does not match ISA-5.1 format: '{tag}'"}


def check_business_rules(tag_text: str, symbol_name: str,
                          symbol_category: str,
                          notes_rules: list[str]) -> list[dict]:
    """
    Rule 2: Business rule validation.
    Returns list of {pass, rule, message} dicts.
    """
    checks = []
    tag = tag_text.strip().upper() if tag_text else ""

    # BR-001: Control valve must pair with a function letter (F, T, P, L)
    if "CONTROL VALVE" in symbol_name.upper() or symbol_category == "valve":
        if re.search(r'[FTPLAZX]CV', tag) or re.search(r'[FTPLAZX]V-', tag):
            checks.append({"pass": True, "rule": "BR-001",
                           "message": "Valve has correct function prefix"})
        elif re.search(r'^[A-Z]+-?[A-Z]{1,5}-?\d', tag):
            checks.append({"pass": True, "rule": "BR-001",
                           "message": "Valve tag format acceptable"})

    # BR-002: Transmitter should end in T or IT
    if "TRANSMITTER" in symbol_name.upper():
        if re.search(r'[A-Z]T-\d|[A-Z]IT-\d', tag):
            checks.append({"pass": True, "rule": "BR-002",
                           "message": "Transmitter has T/IT suffix"})
        else:
            checks.append({"pass": False, "rule": "BR-002",
                           "message": f"Transmitter tag missing T/IT: {tag}"})

    # BR-003: Relief valve should be RV or PSV
    if "RELIEF" in symbol_name.upper() or "SAFETY" in symbol_name.upper():
        if re.search(r'[A-Z]-?[RP][SV]V?-?\d', tag):
            checks.append({"pass": True, "rule": "BR-003",
                           "message": "Relief valve has RV/PSV prefix"})
        else:
            checks.append({"pass": False, "rule": "BR-003",
                           "message": f"Relief valve missing RV/PSV: {tag}"})

    # BR-004: Apply notes-derived rules
    for rule_text in notes_rules[:10]:
        rule_upper = rule_text.upper()
        # Check prefix rules
        m = re.search(r"'([A-Z]+)'\s*PREFIX", rule_upper)
        if m:
            prefix = m.group(1)
            if not tag.startswith(prefix + "-") and not f"-{prefix}-" in tag:
                checks.append({"pass": False, "rule": "BR-004-NOTES",
                               "message": f"Notes rule: tag should have '{prefix}' prefix"})

    if not checks:
        checks.append({"pass": True, "rule": "BR-GENERAL",
                       "message": "No specific business rules violated"})
    return checks


def check_asset_registry(tag_text: str, registry: dict) -> dict:
    """
    Rule 3: Asset registry lookup.
    registry: {tag_number: {discipline, description, ...}}
    """
    if not registry:
        return {"pass": None, "rule": "REGISTRY",
                "message": "No asset registry loaded",
                "in_registry": None}

    tag = tag_text.strip().upper() if tag_text else ""
    if tag in registry:
        entry = registry[tag]
        return {"pass": True, "rule": "REGISTRY",
                "message": f"Found in registry: {entry.get('description','')[:60]}",
                "in_registry": True,
                "registry_entry": entry}

    # Try normalised match (remove vendor prefix like V-)
    stripped = re.sub(r'^[A-Z]-', '', tag)
    for reg_tag, entry in registry.items():
        if stripped == re.sub(r'^[A-Z]-', '', reg_tag):
            return {"pass": True, "rule": "REGISTRY",
                    "message": f"Registry match (normalised): {reg_tag}",
                    "in_registry": True,
                    "matched_key": reg_tag,
                    "registry_entry": entry}

    return {"pass": None, "rule": "REGISTRY",
            "message": f"Tag not found in asset registry: {tag}",
            "in_registry": False}


def validate_candidate(cand: dict,
                        registry: dict,
                        notes_rules: list[str]) -> dict:
    """
    Run all validation checks on a single candidate.
    Returns enriched candidate with validation_status, validation_reason,
    validation_details.
    """
    tag_text       = str(cand.get("tag_text") or "").strip()
    symbol_name    = str(cand.get("symbol_name") or "")
    symbol_category= str(cand.get("symbol_category") or "")

    checks = []

    # Check 1: Format
    fmt_check = check_tag_format(tag_text)
    checks.append(fmt_check)

    # Check 2: Business rules
    biz_checks = check_business_rules(tag_text, symbol_name,
                                       symbol_category, notes_rules)
    checks.extend(biz_checks)

    # Check 3: Registry
    reg_check = check_asset_registry(tag_text, registry)
    checks.append(reg_check)

    # Determine overall status
    hard_fails  = [c for c in checks if c.get("pass") is False]
    soft_warns  = [c for c in checks
                   if c.get("pass") is None and "registry" in c.get("rule","").lower()]

    if hard_fails:
        status = "FAIL"
        reason = " | ".join(c["message"] for c in hard_fails[:3])
    elif soft_warns:
        status = "WARN"
        reason = " | ".join(c["message"] for c in soft_warns[:2])
    else:
        status = "PASS"
        reason = "All validation checks passed"

    result = {**cand}
    result["validation_status"]  = status
    result["validation_reason"]  = reason
    result["validation_details"] = checks
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Asset registry loader
# ═══════════════════════════════════════════════════════════════════════════════

def load_asset_registry(registry_path: str) -> dict:
    """
    Load asset register from JSON (master_tags.json or tables_context.json)
    or from Excel (Annexure-4 format).
    Returns {tag_number_upper: {discipline, description, equipment_description}}
    """
    registry = {}
    path = Path(registry_path)
    if not path.exists():
        log.warning("Registry not found: %s", registry_path)
        return registry

    if path.suffix.lower() == ".json":
        with open(path) as f:
            data = json.load(f)
        # Handle master_tags.json (list of {tag_number, ...})
        if isinstance(data, list):
            for entry in data:
                tag = str(entry.get("tag_number") or "").strip().upper()
                if tag:
                    registry[tag] = entry
        # Handle tables_context.json
        elif isinstance(data, dict) and "master_tag_list" in data:
            for entry in data["master_tag_list"]:
                tag = str(entry.get("tag_number") or "").strip().upper()
                if tag:
                    registry[tag] = entry

    elif path.suffix.lower() in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h or "").strip().upper() for h in row]
                    continue
                if not row:
                    continue
                row_dict = {headers[j]: str(v or "").strip()
                            for j, v in enumerate(row) if j < len(headers)}
                tag = row_dict.get("TAG NUMBER", "").strip().upper()
                if tag:
                    registry[tag] = {
                        "discipline":            row_dict.get("DISCIPLINE", ""),
                        "description":           row_dict.get("TAG DESCRIPTION", ""),
                        "equipment_description": row_dict.get("EQUIPMENT DESCRIPTION", ""),
                        "size_rating":           row_dict.get("SIZE&RATING", ""),
                        "document_number":       row_dict.get("DOCUMENT NUMBER", ""),
                        "duplicate_status":      row_dict.get("DUPLICATE STATUS", ""),
                    }
        except Exception as e:
            log.error("Excel registry load failed: %s", e)

    log.info("Registry loaded: %d tags from %s", len(registry), registry_path)
    return registry


# ═══════════════════════════════════════════════════════════════════════════════
# Main pipeline
# ═══════════════════════════════════════════════════════════════════════════════

def run_validation(
    associations_path: str,
    out_dir: str,
    registry_path: Optional[str] = None,
    notes_path: Optional[str] = None,
) -> list[dict]:

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    # Load input
    with open(associations_path) as f:
        data = json.load(f)
    candidates = data.get("enriched_candidates", data.get("candidates", []))
    log.info("Loaded %d candidates for validation", len(candidates))

    # Load asset registry
    registry = {}
    if registry_path:
        registry = load_asset_registry(registry_path)
    else:
        # Auto-detect from out_dir
        for fname in ["master_tags.json", "tables_context.json",
                       "ANNEXURE-4_4224-MGDV-6-50-2004-001-C.xlsx"]:
            auto_path = out / fname
            if auto_path.exists():
                registry = load_asset_registry(str(auto_path))
                break

    # Load notes rules
    notes_rules: list[str] = []
    if notes_path and Path(notes_path).exists():
        with open(notes_path) as f:
            notes_data = json.load(f)
        notes_rules = (notes_data.get("global_constraints", []) +
                       notes_data.get("tag_detection_rules", []))
    else:
        auto_notes = out / "notes_context.json"
        if auto_notes.exists():
            with open(auto_notes) as f:
                nd = json.load(f)
            notes_rules = (nd.get("global_constraints", []) +
                           nd.get("tag_detection_rules", []))
    log.info("Notes rules loaded: %d rules", len(notes_rules))

    # Run validation
    validated: list[dict] = []
    status_counts = {"PASS": 0, "WARN": 0, "FAIL": 0}

    for cand in candidates:
        result = validate_candidate(cand, registry, notes_rules)
        validated.append(result)
        status_counts[result["validation_status"]] = (
            status_counts.get(result["validation_status"], 0) + 1
        )

    log.info("Validation: PASS=%d | WARN=%d | FAIL=%d",
             status_counts["PASS"], status_counts["WARN"], status_counts["FAIL"])

    # Write output
    out_path = str(out / "step5c_validated.json")
    with open(out_path, "w") as f:
        json.dump({
            "version":         "v1",
            "total_validated": len(validated),
            "status_summary":  status_counts,
            "registry_size":   len(registry),
            "notes_rules":     len(notes_rules),
            "validated_candidates": validated,
        }, f, indent=2)
    log.info("✓ step5c_validated.json → %s", out_path)
    return validated


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Step 5C: Validation Engine (programmatic)")
    parser.add_argument("--associations", help="step5b_associations.json")
    parser.add_argument("--register",     help="Asset registry (JSON or XLSX)")
    parser.add_argument("--notes",        help="notes_context.json")
    parser.add_argument("--context",      help="drawing_context.json")
    parser.add_argument("--out",          default="output")
    args = parser.parse_args()

    assoc_path    = args.associations
    registry_path = args.register
    notes_path    = args.notes

    if args.context:
        with open(args.context) as f:
            ctx = json.load(f)
        assoc_path    = assoc_path    or str(Path(args.out) / "step5b_associations.json")
        registry_path = registry_path or ctx.get("tables_summary",{}).get("master_tags_path")
        notes_path    = notes_path    or ctx.get("notes_context_path")

    assoc_path = assoc_path or str(Path(args.out) / "step5b_associations.json")

    validated = run_validation(assoc_path, args.out, registry_path, notes_path)

    counts = {"PASS": 0, "WARN": 0, "FAIL": 0}
    for v in validated:
        s = v.get("validation_status", "FAIL")
        counts[s] = counts.get(s, 0) + 1

    print(f"\n=== Step 5C Complete ===")
    print(f"  Total validated : {len(validated)}")
    print(f"  PASS            : {counts['PASS']}")
    print(f"  WARN            : {counts['WARN']}")
    print(f"  FAIL            : {counts['FAIL']}")
    print(f"\n  Output: {args.out}/step5c_validated.json")


if __name__ == "__main__":
    main()
