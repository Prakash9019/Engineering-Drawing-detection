"""
Stage 7: Excel Output Generation (Architecture Layer 16)
==========================================================
Generates the final structured engineering tag register as an Excel workbook
with the following sheets:

  Sheet 1 — Tag Register   (15-field standard format)
  Sheet 2 — Notes          (engineering notes with classification)
  Sheet 3 — Metadata       (title block fields)
  Sheet 4 — Summary        (counts by discipline, confidence routing)
  Sheet 5 — Cloud Scope    (revision cloud info)
"""
import logging
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from settings import EXCEL_HEADERS, EXCEL_COL_WIDTHS

log = logging.getLogger(__name__)


def _style_header(cell):
    cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E79')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        top=Side('thin'), bottom=Side('thin'),
        left=Side('thin'), right=Side('thin'),
    )


def _style_data(cell, alt_row: bool):
    cell.font = Font(name='Arial', size=9)
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    cell.border = Border(
        top=Side('thin'), bottom=Side('thin'),
        left=Side('thin'), right=Side('thin'),
    )
    if alt_row:
        cell.fill = PatternFill('solid', fgColor='E8F0FE')


def _write_register_sheet(ws, records: List[dict], title_block: dict):
    """Sheet 1: main tag register."""
    doc_num = title_block.get('document_number', '')
    sheet_no = title_block.get('sheet_number', '001')
    rev = title_block.get('revision', '')
    dwg_ref = title_block.get('drawing_number', '') or doc_num
    doc_title = title_block.get('drawing_title', '')
    status = title_block.get('status', '')
    date_str = title_block.get('date', '')

    # Header row
    for c, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=header)
        _style_header(cell)

    # Data rows
    for i, rec in enumerate(records, 1):
        row_data = [
            i,
            rec.get('discipline', ''),
            rec.get('tag_number', ''),
            rec.get('tag_description', ''),
            rec.get('equipment_description', ''),
            rec.get('size_rating', ''),
            doc_num,
            sheet_no,
            rev,
            dwg_ref,
            doc_title,
            status,
            date_str,
            rec.get('duplicate', 'NO'),
            rec.get('remarks', ''),
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=c, value=val)
            _style_data(cell, alt_row=(i % 2 == 0))

    # Column widths
    for i, w in enumerate(EXCEL_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze + filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:O{len(records) + 1}"


def _write_notes_sheet(ws, notes_data: dict):
    """Sheet 2: engineering notes."""
    notes = notes_data.get('notes', []) if isinstance(notes_data, dict) else []
    if not notes:
        return
    headers = ["NOTE #", "TEXT", "TYPE"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, value=h)
        _style_header(cell)
    for i, n in enumerate(notes, 2):
        ws.cell(i, 1, value=n.get('note_number', ''))
        ws.cell(i, 2, value=n.get('text', ''))
        ws.cell(i, 3, value=n.get('type', ''))
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 18


def _write_metadata_sheet(ws, title_block: dict):
    """Sheet 3: title block metadata."""
    ws.cell(1, 1, value="FIELD").font = Font(bold=True)
    ws.cell(1, 2, value="VALUE").font = Font(bold=True)
    for i, (k, v) in enumerate(title_block.items(), 2):
        ws.cell(i, 1, value=k.upper().replace('_', ' '))
        ws.cell(i, 2, value=str(v))
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def _write_summary_sheet(ws, records: List[dict], title_block: dict, cloud_info: dict):
    """Sheet 4: summary statistics."""
    ws['A1'] = "CDCI TAG REGISTER — EXTRACTION SUMMARY"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws['A3'] = "Drawing:"
    ws['B3'] = title_block.get('drawing_number', '')
    ws['A4'] = "Title:"
    ws['B4'] = title_block.get('drawing_title', '')
    ws['A5'] = "Revision:"
    ws['B5'] = title_block.get('revision', '')
    ws['A6'] = "Total Tags:"
    ws['B6'] = len(records)

    # Routing summary
    routes = {}
    for r in records:
        routes[r.get('route', 'UNKNOWN')] = routes.get(r.get('route', 'UNKNOWN'), 0) + 1
    ws['A8'] = "AUTO_ACCEPT:"
    ws['B8'] = routes.get('AUTO_ACCEPT', 0)
    ws['A9'] = "REVIEW_REQUIRED:"
    ws['B9'] = routes.get('REVIEW_REQUIRED', 0)
    ws['A10'] = "AUTO_REJECT:"
    ws['B10'] = routes.get('AUTO_REJECT', 0)

    # Cloud scope info
    ws['A12'] = "CLOUD SCOPE:"
    if cloud_info.get('is_full_scope'):
        ws['B12'] = "FULL (no clouds detected)"
    else:
        ws['B12'] = f"{cloud_info.get('num_clouds', 0)} clouds, " \
                    f"{cloud_info.get('coverage_pct', 0):.1f}% coverage"

    # By discipline
    ws['A14'] = "DISCIPLINE"
    ws['B14'] = "COUNT"
    ws['A14'].font = Font(bold=True)
    ws['B14'].font = Font(bold=True)
    disc_counts = {}
    for r in records:
        d = r.get('discipline', '?')
        disc_counts[d] = disc_counts.get(d, 0) + 1
    for i, (k, v) in enumerate(sorted(disc_counts.items()), 15):
        ws.cell(i, 1, value=k)
        ws.cell(i, 2, value=v)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def _write_cloud_sheet(ws, cloud_info: dict):
    """Sheet 5: revision cloud details."""
    ws.cell(1, 1, value="CLOUD #").font = Font(bold=True)
    ws.cell(1, 2, value="BOUNDING BOX (x0,y0,x1,y1)").font = Font(bold=True)
    ws.cell(1, 3, value="POLYGON VERTICES").font = Font(bold=True)
    for i, bb in enumerate(cloud_info.get('bounding_boxes', []), 2):
        ws.cell(i, 1, value=i - 1)
        ws.cell(i, 2, value=str(bb))
        poly_count = len(cloud_info.get('polygons', [])[i - 2]) if cloud_info.get('polygons') else 0
        ws.cell(i, 3, value=poly_count)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20


def generate_excel(
    records: List[dict],
    title_block: dict,
    notes_data: dict,
    cloud_info: dict,
    output_path: Path,
):
    """
    Generate the final Excel workbook.

    Args:
        records: Validated tag records
        title_block: Drawing metadata
        notes_data: Notes from stage 3
        cloud_info: Cloud detection result as dict
        output_path: Path to write the .xlsx file
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    sheet_name = (title_block.get('document_number', '') or 'TAG_REGISTER')[:31]
    ws1.title = sheet_name

    _write_register_sheet(ws1, records, title_block)

    notes = notes_data.get('notes', []) if isinstance(notes_data, dict) else []
    if notes:
        _write_notes_sheet(wb.create_sheet("Notes"), notes_data)

    _write_metadata_sheet(wb.create_sheet("Metadata"), title_block)
    _write_summary_sheet(wb.create_sheet("Summary"), records, title_block, cloud_info)

    if cloud_info.get('bounding_boxes'):
        _write_cloud_sheet(wb.create_sheet("Cloud Scope"), cloud_info)

    wb.save(str(output_path))
    log.info(f"  Excel written: {output_path} ({len(records)} records, "
             f"{len(wb.sheetnames)} sheets)")
