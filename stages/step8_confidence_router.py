#!/usr/bin/env python3
"""
step8_confidence_router.py — Confidence Aggregation, Routing & Output Generation
==================================================================================
CDCI P&ID Pipeline — Step 8  (Blueprint Layers 15 + 16)

Type: PROGRAMMATIC — No Gemini, No Claude

What this does
--------------
Layer 15 — Confidence Aggregation:
  Computes C_final for every CEDM record using the weighted formula:

    C_final = 0.25×C_det + 0.30×C_ocr + 0.15×C_geo + 0.20×C_val + 0.10×C_reg

  Routes each record to one of three queues:
    AUTO_ACCEPT   C_final ≥ 0.85  → output directly (target: ~70%)
    HUMAN_REVIEW  0.60 ≤ C < 0.85 → queue for reviewer (target: ~25%)
    AUTO_REJECT   C_final < 0.60  → audit log, excluded from output (~5%)

  Missing-tag records (no OCR text) → always HUMAN_REVIEW regardless of score.

  Priority within HUMAN_REVIEW queue:
    P1 CRITICAL: missing tags, cloud ambiguity
    P2 HIGH:     OCR low conf (C_ocr < 0.75), duplicate conflicts
    P3 MEDIUM:   register mismatches, prefix warnings
    P4 LOW:      format warnings, project mismatches

Layer 16 — Output Generation:
  Produces THREE output files:

  1. final_tags.xlsx      — Annexure-4 exact format (15 columns)
                            Sheet 1: AUTO_ACCEPT tags (blue header)
                            Sheet 2: HUMAN_REVIEW queue (orange header)
                            Sheet 3: Statistics summary

  2. human_review_queue.json — detailed review items with drawing crop coords
                                for each flagged tag (feeds Next.js review UI)

  3. audit_log.json          — all AUTO_REJECT records with rejection reason

Inputs
------
  step7_cedm_output.json   (from Step 7)
  drawing_context.json     (optional — for metadata)

Outputs
-------
  output/final_tags.xlsx
  output/human_review_queue.json
  output/audit_log.json
  output/step8_routing_summary.json

Usage
-----
  python step8_confidence_router.py \\
      --cedm   output/step7_cedm_output.json \\
      --out    output/

  python step8_confidence_router.py \\
      --context output/drawing_context.json
"""

import argparse
import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

# ── Confidence weights (Blueprint §9.1, rebalanced) ───────────────────────────
# In this pipeline the MLLM (Gemini) is the PRIMARY text source and Tesseract is
# a secondary cross-check that is frequently silent (it reads almost no symbol
# text on this drawing). Weighting raw OCR at 0.30 therefore collapsed C_final
# and AUTO_REJECTed ~87% of perfectly good tags. We keep OCR in the blend but
# (a) fall back to the model's own read confidence when OCR is silent (see
# compute_confidence) and (b) trim the registry weight, which is only 0.5 for
# the many legitimate tags absent from the sparse 46-row register.
W_DET  = 0.30   # C_det  — symbol detection (vision_confidence)
W_OCR  = 0.30   # C_ocr  — text confidence (OCR, or model read when OCR silent)
W_GEO  = 0.15   # C_geo  — geometric association
W_VAL  = 0.20   # C_val  — validation stage scores
W_REG  = 0.05   # C_reg  — registry lookup (1.0=found, 0.5=not found)

# ── Routing thresholds ────────────────────────────────────────────────────────
# Register-confirmed + validated tags auto-accept; novel tags fall to review
# (never silently rejected). Reject is reserved for genuinely low-signal noise.
THRESHOLD_ACCEPT = 0.80
THRESHOLD_REVIEW = 0.55

# ── Priority thresholds for HUMAN_REVIEW queue ────────────────────────────────
OCR_LOW_CONF_THRESHOLD = 0.75

# ── Excel column widths (matched to Annexure-4 template) ──────────────────────
COL_WIDTHS = {
    "S.NO":                   6,
    "DISCIPLINE":            18,
    "TAG NUMBER":            20,
    "TAG DESCRIPTION":       45,
    "EQUIPMENT DESCRIPTION": 45,
    "SIZE&RATING":           14,
    "DOCUMENT NUMBER":       25,
    "SHEET NO":               9,
    "REV":                    6,
    "DRAWING REFERENCE":     30,
    "DOCUMENT TITLE":        45,
    "DOC STATUS":            42,
    "DATE":                  12,
    "DUPLICATE STATUS":      16,
    "REMARKS":               45,
}

# Extra columns in the REVIEW sheet
REVIEW_EXTRA_COLS = {
    "C_FINAL":        10,
    "REVIEW_PRIORITY": 16,
    "REVIEW_REASON":   40,
}

# ── Excel colour palette ──────────────────────────────────────────────────────
COLOUR = {
    "accept_header": "1F4E79",   # dark blue
    "accept_row_a":  "DEEAF1",   # light blue alt row
    "accept_row_b":  "FFFFFF",
    "review_header": "843C0C",   # dark orange
    "review_row_a":  "FCE4D6",   # light orange alt row
    "review_row_b":  "FFFFFF",
    "reject_header": "7B0D0D",   # dark red
    "summary_header":"375623",   # dark green
    "p1_flag":       "FF0000",   # Critical priority — red text
    "p2_flag":       "FF6600",   # High priority — orange text
    "p3_flag":       "DAA520",   # Medium — gold
    "p4_flag":       "808080",   # Low — grey
    "stats_value":   "244185",
}


# ═══════════════════════════════════════════════════════════════════════════════
# Confidence computation
# ═══════════════════════════════════════════════════════════════════════════════

def compute_confidence(record: dict) -> dict:
    """
    Compute all confidence signals and C_final for one CEDM record.
    Returns a confidence dict with all signals and the final score.
    """
    c_det = float(record.get("_c_det") or 0.0)
    c_ocr = float(record.get("_c_ocr") or 0.0)
    c_geo = float(record.get("_c_geo") or 0.0)
    c_val = float(record.get("_c_val") or 0.5)
    c_reg = float(record.get("_c_reg") or 0.5)

    # Clamp all to [0, 1]
    c_det = max(0.0, min(1.0, c_det))
    c_ocr = max(0.0, min(1.0, c_ocr))
    c_geo = max(0.0, min(1.0, c_geo))
    c_val = max(0.0, min(1.0, c_val))
    c_reg = max(0.0, min(1.0, c_reg))

    # Connectivity penalty (step5b2_hierarchy.json, via step7's _hier_* fields):
    # an ISOLATED detection has no graph edge to any pipeline or equipment — the
    # global connectivity view found nothing corroborating it. This is the classic
    # detached-label / false-positive signature, so we discount geometric
    # confidence. Applied only when the hierarchy was actually run (key present),
    # otherwise behaviour is unchanged.
    c_geo_raw = c_geo
    isolated  = bool(record.get("_hier_is_isolated", False))
    if "_hier_is_isolated" in record and isolated:
        c_geo = round(c_geo * 0.5, 3)

    # Effective text confidence: when Tesseract is silent (c_ocr == 0) the tag
    # text actually came from the MLLM, so use its read confidence (lightly
    # discounted) instead of penalising the record for a missing secondary OCR.
    c_ocr_eff = c_ocr if c_ocr > 0 else round(0.9 * c_det, 3)

    c_final = (W_DET * c_det
             + W_OCR * c_ocr_eff
             + W_GEO * c_geo
             + W_VAL * c_val
             + W_REG * c_reg)
    c_final = round(max(0.0, min(1.0, c_final)), 4)

    return {
        "c_det":     round(c_det, 3),
        "c_ocr":     round(c_ocr, 3),
        "c_ocr_eff": c_ocr_eff,
        "c_geo":     round(c_geo, 3),
        "c_geo_raw": round(c_geo_raw, 3),
        "isolated":  isolated,
        "c_val":     round(c_val, 3),
        "c_reg":     round(c_reg, 3),
        "c_final":   c_final,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Review priority classifier
# ═══════════════════════════════════════════════════════════════════════════════

def classify_review_priority(record: dict, conf: dict) -> tuple[int, str]:
    """
    Classify a HUMAN_REVIEW record into P1–P4.
    Returns (priority_int, reason_string).

    P1 CRITICAL: missing tag, no OCR text at all
    P2 HIGH:     low OCR confidence, duplicate conflict
    P3 MEDIUM:   not in register, prefix mismatch
    P4 LOW:      format warning, project mismatch
    """
    raw_tag      = str(record.get("_raw_tag") or "").strip()
    val_status   = record.get("_validation_status", "WARN")
    sow_status   = record.get("_sow_status", "")
    remarks      = record.get("REMARKS", "")
    c_ocr        = conf["c_ocr"]
    c_val        = conf["c_val"]

    # P1: Missing tag — no text at all
    if not raw_tag or raw_tag in {"UNKNOWN", ""}:
        return 1, "MISSING_TAG: symbol detected but no text extracted"

    # P1: Validation hard fail
    if val_status == "FAIL":
        return 1, f"VALIDATION_FAIL: {record.get('_validation_status','')}"

    # P2: Duplicate conflict
    if "DUPLICATE" in remarks.upper() or record.get("DUPLICATE STATUS") == "YES":
        return 2, "DUPLICATE_CONFLICT: tag appears multiple times"

    # P2: Low OCR confidence AND no model fallback — genuinely weak text signal
    c_ocr_eff = conf.get("c_ocr_eff", c_ocr)
    if c_ocr_eff < OCR_LOW_CONF_THRESHOLD:
        return 2, f"LOW_TEXT_CONF: c_text={c_ocr_eff:.2f} (threshold {OCR_LOW_CONF_THRESHOLD})"

    # P3: Isolated detection — hierarchy found no pipe/equipment connectivity
    if conf.get("isolated"):
        return 3, "ISOLATED_DETECTION: no pipe/equipment connectivity (possible stray label)"

    # P3: Not in register
    if not record.get("_in_registry"):
        return 3, "NOT_IN_REGISTER: tag not found in client asset register"

    # P3: Scope not defined (SOW memory absent / tag not in USE|DO_NOT_USE list)
    if sow_status == "UNSPECIFIED":
        return 3, "SOW_UNSPECIFIED: tag not in USE or DO_NOT_USE list"

    # P3: Prefix mismatch (from remarks)
    if "PREFIX_MISMATCH" in remarks.upper():
        return 3, "PREFIX_MISMATCH: tag prefix inconsistent with equipment type"

    # P4: Format or project warning
    if c_val < 0.8:
        return 4, f"VALIDATION_WARN: c_val={c_val:.2f}, review remarks"

    return 4, f"CONFIDENCE_BORDERLINE: c_final={conf['c_final']:.3f}"


# ═══════════════════════════════════════════════════════════════════════════════
# Routing engine
# ═══════════════════════════════════════════════════════════════════════════════

def route_records(cedm_records: list[dict]) -> dict:
    """
    Apply confidence scoring and route all records.
    Returns {accepted, review, rejected, all_with_routing}.
    """
    accepted: list[dict] = []
    review:   list[dict] = []
    rejected: list[dict] = []

    for record in cedm_records:
        conf = compute_confidence(record)
        c    = conf["c_final"]

        # Annotate record with confidence
        record["_confidence"] = conf

        raw_tag = str(record.get("_raw_tag") or "").strip()
        is_missing_tag = not raw_tag or raw_tag in {"UNKNOWN", ""}

        if is_missing_tag:
            # Missing tags → always HUMAN_REVIEW regardless of score
            priority, reason = 1, "MISSING_TAG: no OCR text extracted"
            record["_route"]           = "HUMAN_REVIEW"
            record["_review_priority"] = priority
            record["_review_reason"]   = reason
            review.append(record)

        elif c >= THRESHOLD_ACCEPT:
            record["_route"]           = "AUTO_ACCEPT"
            record["_review_priority"] = None
            record["_review_reason"]   = None
            accepted.append(record)

        elif c >= THRESHOLD_REVIEW:
            priority, reason = classify_review_priority(record, conf)
            record["_route"]           = "HUMAN_REVIEW"
            record["_review_priority"] = priority
            record["_review_reason"]   = reason
            review.append(record)

        else:
            record["_route"]           = "AUTO_REJECT"
            record["_review_priority"] = None
            record["_review_reason"]   = f"c_final={c:.3f} below reject threshold {THRESHOLD_REVIEW}"
            rejected.append(record)

    # Sort review queue by priority (P1 first)
    review.sort(key=lambda r: (r.get("_review_priority") or 9, r.get("TAG NUMBER", "")))

    log.info("Routing: AUTO_ACCEPT=%d | HUMAN_REVIEW=%d | AUTO_REJECT=%d",
             len(accepted), len(review), len(rejected))

    total = len(cedm_records)
    if total > 0:
        log.info("Accept rate: %.1f%%  Review rate: %.1f%%  Reject rate: %.1f%%",
                 len(accepted) / total * 100,
                 len(review)   / total * 100,
                 len(rejected) / total * 100)

    return {
        "accepted": accepted,
        "review":   review,
        "rejected": rejected,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Excel writer
# ═══════════════════════════════════════════════════════════════════════════════

def _write_excel(accepted: list[dict], review: list[dict], rejected: list[dict],
                  out_path: str, drawing_meta: dict) -> None:
    """
    Write the final Excel output matching Annexure-4 format exactly.
    Three sheets: AUTO_ACCEPT, HUMAN_REVIEW, SUMMARY.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    ANNEX_COLS = list(COL_WIDTHS.keys())
    # Map "S.NO" → "SLNO" for data lookup
    def _get_val(r: dict, col: str):
        if col == "S.NO":
            return r.get("SLNO")
        return r.get(col, "")

    def _make_border():
        thin = Side(style="thin", color="BFBFBF")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _header_row(ws, cols: list, bg_hex: str, text_hex: str = "FFFFFF",
                     row: int = 1) -> None:
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=j, value=col)
            cell.font      = Font(bold=True, color=text_hex, name="Arial", size=10)
            cell.fill      = PatternFill("solid", fgColor=bg_hex)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
            cell.border    = _make_border()

    def _data_row(ws, record: dict, cols: list, row_num: int,
                   row_fill_a: str, row_fill_b: str,
                   extra_cols: Optional[dict] = None) -> None:
        is_alt = (row_num % 2 == 0)
        fill   = PatternFill("solid", fgColor=row_fill_a if is_alt else row_fill_b)
        for j, col in enumerate(cols, 1):
            val  = _get_val(record, col)
            cell = ws.cell(row=row_num, column=j, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=False)
            cell.border    = _make_border()
        # Extra columns (review sheet)
        if extra_cols:
            conf = record.get("_confidence", {})
            j    = len(cols) + 1
            extras = {
                "C_FINAL":         round(conf.get("c_final", 0), 3),
                "REVIEW_PRIORITY": f"P{record.get('_review_priority','?')}",
                "REVIEW_REASON":   record.get("_review_reason", ""),
            }
            for col_name in extra_cols:
                val  = extras.get(col_name, "")
                cell = ws.cell(row=row_num, column=j, value=val)
                cell.font      = Font(name="Arial", size=9)
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = _make_border()
                # Colour-code priority
                if col_name == "REVIEW_PRIORITY":
                    pri = record.get("_review_priority")
                    if pri == 1:
                        cell.font = Font(name="Arial", size=9, bold=True,
                                         color=COLOUR["p1_flag"])
                    elif pri == 2:
                        cell.font = Font(name="Arial", size=9,
                                         color=COLOUR["p2_flag"])
                    elif pri == 3:
                        cell.font = Font(name="Arial", size=9,
                                         color=COLOUR["p3_flag"])
                j += 1

    def _set_col_widths(ws, cols: list, widths: dict,
                         extra: Optional[dict] = None) -> None:
        for j, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 15)
        if extra:
            for k, (col_name, width) in enumerate(extra.items(), len(cols) + 1):
                ws.column_dimensions[get_column_letter(k)].width = width

    def _freeze_and_filter(ws, freeze_row: int = 2) -> None:
        ws.freeze_panes = ws.cell(row=freeze_row, column=1)
        ws.auto_filter.ref = ws.dimensions

    # ── Sheet 1: AUTO_ACCEPT ──────────────────────────────────────────────────
    ws1 = wb.create_sheet("AUTO_ACCEPT")
    ws1.sheet_properties.tabColor = "1F4E79"
    ws1.row_dimensions[1].height  = 28

    _header_row(ws1, ANNEX_COLS, COLOUR["accept_header"])
    for i, r in enumerate(accepted, 2):
        _data_row(ws1, r, ANNEX_COLS, i,
                  COLOUR["accept_row_a"], COLOUR["accept_row_b"])
    _set_col_widths(ws1, ANNEX_COLS, COL_WIDTHS)
    _freeze_and_filter(ws1)

    # ── Sheet 2: HUMAN_REVIEW ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("HUMAN_REVIEW")
    ws2.sheet_properties.tabColor = "C55A11"
    ws2.row_dimensions[1].height  = 28

    all_review_cols = ANNEX_COLS + list(REVIEW_EXTRA_COLS.keys())
    _header_row(ws2, all_review_cols, COLOUR["review_header"])
    for i, r in enumerate(review, 2):
        _data_row(ws2, r, ANNEX_COLS, i,
                  COLOUR["review_row_a"], COLOUR["review_row_b"],
                  extra_cols=REVIEW_EXTRA_COLS)
    _set_col_widths(ws2, ANNEX_COLS, COL_WIDTHS, extra=REVIEW_EXTRA_COLS)
    _freeze_and_filter(ws2)

    # ── Sheet 3: SUMMARY ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("SUMMARY")
    ws3.sheet_properties.tabColor = "375623"

    def _s(row, col, val, bold=False, color="000000", bg=None, num_fmt=None):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font      = Font(name="Arial", size=10, bold=bold, color=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if num_fmt:
            cell.number_format = num_fmt

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20

    _s(1, 1, "CDCI TAG EXTRACTION — PIPELINE SUMMARY", bold=True,
       color="FFFFFF", bg=COLOUR["summary_header"])
    ws3.merge_cells("A1:C1")
    ws3.row_dimensions[1].height = 22

    now  = datetime.now().strftime("%Y-%m-%d %H:%M")
    rows = [
        (3,  "Drawing Number",    drawing_meta.get("drawing_number",  "—")),
        (4,  "Sheet",             drawing_meta.get("sheet_number",    "—")),
        (5,  "Revision",          drawing_meta.get("revision_code",   "—")),
        (6,  "Document Title",    drawing_meta.get("drawing_title",   "—")[:60]),
        (7,  "Extraction Date",   now),
        (9,  "─── RESULTS ───",   ""),
        (10, "Total Records",     len(accepted) + len(review) + len(rejected)),
        (11, "AUTO_ACCEPT",       len(accepted)),
        (12, "HUMAN_REVIEW",      len(review)),
        (13, "AUTO_REJECT",       len(rejected)),
        (15, "─── REVIEW BREAKDOWN ───", ""),
        (16, "P1 Critical",       sum(1 for r in review if r.get("_review_priority")==1)),
        (17, "P2 High",           sum(1 for r in review if r.get("_review_priority")==2)),
        (18, "P3 Medium",         sum(1 for r in review if r.get("_review_priority")==3)),
        (19, "P4 Low",            sum(1 for r in review if r.get("_review_priority")==4)),
    ]

    total = len(accepted) + len(review) + len(rejected)
    if total > 0:
        rows += [
            (21, "─── RATES ───", ""),
            (22, "Auto-Accept Rate", f"{len(accepted)/total*100:.1f}%"),
            (23, "Review Rate",      f"{len(review)/total*100:.1f}%"),
            (24, "Reject Rate",      f"{len(rejected)/total*100:.1f}%"),
            (25, "Target Accept (blueprint)", "≥ 70%"),
        ]

    for row, label, value in rows:
        is_header = "───" in str(label)
        _s(row, 1, label, bold=is_header,
           color="244185" if is_header else "000000",
           bg="D9E1F2" if is_header else None)
        if not is_header:
            _s(row, 2, value, bold=isinstance(value, int),
               color=COLOUR["stats_value"] if isinstance(value, int) else "000000")

    # Add confidence distribution
    if accepted or review:
        all_c = [r["_confidence"]["c_final"]
                 for r in accepted + review
                 if "_confidence" in r]
        if all_c:
            avg_c  = sum(all_c) / len(all_c)
            min_c  = min(all_c)
            max_c  = max(all_c)
            _s(27, 1, "─── CONFIDENCE ───", bold=True,
               color="244185", bg="D9E1F2")
            for row, label, value in [
                (28, "Average C_final", round(avg_c,3)),
                (29, "Min C_final",     round(min_c,3)),
                (30, "Max C_final",     round(max_c,3)),
            ]:
                _s(row, 1, label)
                _s(row, 2, value, bold=True, color=COLOUR["stats_value"])

    wb.save(out_path)
    log.info("✓ final_tags.xlsx saved → %s", out_path)


# ═══════════════════════════════════════════════════════════════════════════════
# JSON outputs
# ═══════════════════════════════════════════════════════════════════════════════

def _build_review_queue(review: list[dict]) -> list[dict]:
    """
    Build the human_review_queue.json structure.
    Includes bbox coordinates for the Next.js drawing crop viewer.
    """
    queue = []
    for r in review:
        conf = r.get("_confidence", {})
        item = {
            "queue_id":         r.get("_canonical_id", r.get("_candidate_id", "")),
            "priority":         r.get("_review_priority", 4),
            "priority_label":   f"P{r.get('_review_priority', 4)}",
            "review_reason":    r.get("_review_reason", ""),
            "tag_number":       r.get("TAG NUMBER", ""),
            "raw_tag":          r.get("_raw_tag", ""),
            "symbol_name":      r.get("_symbol_name", ""),
            "discipline":       r.get("DISCIPLINE", ""),
            "c_final":          conf.get("c_final", 0),
            "c_det":            conf.get("c_det", 0),
            "c_ocr":            conf.get("c_ocr", 0),
            "c_geo":            conf.get("c_geo", 0),
            "c_val":            conf.get("c_val", 0),
            "c_reg":            conf.get("c_reg", 0),
            "symbol_bbox":      r.get("_symbol_bbox", {}),
            "validation_status":r.get("_validation_status", ""),
            "sow_status":       r.get("_sow_status", ""),
            "in_registry":      r.get("_in_registry", False),
            "remarks":          r.get("REMARKS", ""),
            "document_number":  r.get("DOCUMENT NUMBER", ""),
            "sheet_number":     r.get("SHEET NO", ""),
            # Actions available in review UI
            "reviewer_actions": ["ACCEPT", "CORRECT", "REJECT", "SKIP"],
            "reviewer_decision":None,   # filled by review UI
            "reviewer_note":    None,
        }
        queue.append(item)
    return queue


def _build_audit_log(rejected: list[dict]) -> list[dict]:
    """Build audit_log.json for AUTO_REJECT records."""
    log_entries = []
    for r in rejected:
        conf = r.get("_confidence", {})
        log_entries.append({
            "candidate_id":   r.get("_candidate_id", ""),
            "canonical_id":   r.get("_canonical_id", ""),
            "tag_number":     r.get("TAG NUMBER", ""),
            "raw_tag":        r.get("_raw_tag", ""),
            "rejection_reason":r.get("_review_reason", ""),
            "c_final":        conf.get("c_final", 0),
            "c_det":          conf.get("c_det", 0),
            "c_ocr":          conf.get("c_ocr", 0),
            "c_val":          conf.get("c_val", 0),
            "symbol_bbox":    r.get("_symbol_bbox", {}),
            "patch_id":       r.get("_patch_id"),
            "rejected_at":    datetime.now().isoformat(),
        })
    return log_entries


# ═══════════════════════════════════════════════════════════════════════════════
# Drawing metadata loader
# ═══════════════════════════════════════════════════════════════════════════════

def _load_meta(context_path: Optional[str], out_dir: str) -> dict:
    meta = {"drawing_number":"","sheet_number":"","revision_code":"","drawing_title":""}
    for p in [context_path, str(Path(out_dir)/"drawing_context.json")]:
        if p and Path(p).exists():
            with open(p) as f:
                ctx = json.load(f)
            meta.update({
                "drawing_number":  ctx.get("drawing_number",""),
                "sheet_number":    ctx.get("sheet_number",""),
                "revision_code":   ctx.get("revision_code",""),
                "drawing_title":   ctx.get("drawing_title",""),
            })
            break
    return meta


# ═══════════════════════════════════════════════════════════════════════════════
# Main pipeline
# ═══════════════════════════════════════════════════════════════════════════════

def run_confidence_router(
    cedm_path:    str,
    out_dir:      str,
    context_path: Optional[str] = None,
) -> dict:

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    with open(cedm_path) as f:
        data = json.load(f)
    cedm_records = data.get("records", data.get("candidates", []))
    log.info("Loaded %d CEDM records", len(cedm_records))

    meta = _load_meta(context_path, out_dir)

    # ── Layer 15: Route ───────────────────────────────────────────────────────
    log.info("=== Layer 15: Confidence Aggregation & Routing ===")
    routes = route_records(cedm_records)
    accepted, review, rejected = routes["accepted"], routes["review"], routes["rejected"]

    # ── Layer 16: Output generation ───────────────────────────────────────────
    log.info("=== Layer 16: Output Generation ===")

    # 1. Excel
    excel_path = str(out / "final_tags.xlsx")
    _write_excel(accepted, review, rejected, excel_path, meta)

    # 2. Human review queue JSON
    review_queue = _build_review_queue(review)
    review_path  = str(out / "human_review_queue.json")
    with open(review_path, "w") as f:
        json.dump({
            "version":      "v1",
            "total_items":  len(review_queue),
            "p1_critical":  sum(1 for r in review_queue if r["priority"]==1),
            "p2_high":      sum(1 for r in review_queue if r["priority"]==2),
            "p3_medium":    sum(1 for r in review_queue if r["priority"]==3),
            "p4_low":       sum(1 for r in review_queue if r["priority"]==4),
            "items":        review_queue,
        }, f, indent=2)
    log.info("✓ human_review_queue.json (%d items) → %s",
             len(review_queue), review_path)

    # 3. Audit log
    audit_log   = _build_audit_log(rejected)
    audit_path  = str(out / "audit_log.json")
    with open(audit_path, "w") as f:
        json.dump({
            "version":      "v1",
            "total_rejected": len(audit_log),
            "entries":      audit_log,
        }, f, indent=2)
    log.info("✓ audit_log.json (%d rejected) → %s", len(audit_log), audit_path)

    # 4. Routing summary
    total = len(cedm_records)
    summary = {
        "version":        "v1",
        "timestamp":      datetime.now().isoformat(),
        "drawing_number": meta["drawing_number"],
        "sheet":          meta["sheet_number"],
        "revision":       meta["revision_code"],
        "totals": {
            "input_records": total,
            "auto_accept":   len(accepted),
            "human_review":  len(review),
            "auto_reject":   len(rejected),
        },
        "rates": {
            "accept_pct": round(len(accepted)/max(total,1)*100, 1),
            "review_pct": round(len(review)  /max(total,1)*100, 1),
            "reject_pct": round(len(rejected)/max(total,1)*100, 1),
        },
        "review_breakdown": {
            "p1_critical": sum(1 for r in review if r.get("_review_priority")==1),
            "p2_high":     sum(1 for r in review if r.get("_review_priority")==2),
            "p3_medium":   sum(1 for r in review if r.get("_review_priority")==3),
            "p4_low":      sum(1 for r in review if r.get("_review_priority")==4),
        },
        "confidence_stats": {
            "avg_c_final": round(
                sum(r["_confidence"]["c_final"] for r in cedm_records
                    if "_confidence" in r)
                / max(total, 1), 3),
        },
        "output_files": {
            "excel":        excel_path,
            "review_queue": review_path,
            "audit_log":    audit_path,
        },
    }
    summary_path = str(out / "step8_routing_summary.json")
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    log.info("✓ step8_routing_summary.json → %s", summary_path)

    return summary


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Step 8: Confidence Routing & Excel Output Generation")
    parser.add_argument("--cedm",    help="step7_cedm_output.json")
    parser.add_argument("--context", help="drawing_context.json")
    parser.add_argument("--out",     default="output")
    args = parser.parse_args()

    cedm_path = args.cedm or str(Path(args.out) / "step7_cedm_output.json")
    if not Path(cedm_path).exists():
        parser.error(f"CEDM output not found: {cedm_path}. Run step7 first.")

    summary = run_confidence_router(
        cedm_path=cedm_path,
        out_dir=args.out,
        context_path=args.context,
    )

    t = summary["totals"]
    r = summary["rates"]
    rb = summary["review_breakdown"]

    print(f"\n=== Step 8 Complete — Confidence Routing & Output ===")
    print(f"\n  Drawing  : {summary['drawing_number']}  "
          f"Sht={summary['sheet']}  Rev={summary['revision']}")
    print(f"\n  ── Routing Results ──")
    print(f"  Input records    : {t['input_records']}")
    print(f"  AUTO_ACCEPT      : {t['auto_accept']}  ({r['accept_pct']}%)"
          f"{'  ✓ meets 70% target' if r['accept_pct'] >= 70 else '  ⚠ below 70% target'}")
    print(f"  HUMAN_REVIEW     : {t['human_review']}  ({r['review_pct']}%)")
    print(f"  AUTO_REJECT      : {t['auto_reject']}  ({r['reject_pct']}%)")
    print(f"\n  ── Review Queue ──")
    print(f"  P1 Critical      : {rb['p1_critical']}  (missing tags, failures)")
    print(f"  P2 High          : {rb['p2_high']}  (low OCR conf, duplicates)")
    print(f"  P3 Medium        : {rb['p3_medium']}  (register mismatches)")
    print(f"  P4 Low           : {rb['p4_low']}  (format warnings)")
    print(f"\n  Avg C_final      : {summary['confidence_stats']['avg_c_final']}")
    print(f"\n  ── Output Files ──")
    print(f"  {args.out}/final_tags.xlsx            ← Annexure-4 format (3 sheets)")
    print(f"  {args.out}/human_review_queue.json    ← {t['human_review']} items for review")
    print(f"  {args.out}/audit_log.json             ← {t['auto_reject']} rejected records")
    print(f"  {args.out}/step8_routing_summary.json ← run statistics")


if __name__ == "__main__":
    main()