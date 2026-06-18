#!/usr/bin/env python3
"""
Compare final_tags.xlsx (AUTO_ACCEPT / HUMAN_REVIEW) vs Annexure-4 register.

Writes output/final_tags_vs_annexure4.xlsx with sheets:
  SUMMARY                  — unique register coverage + duplicate/suspicious flags
  AUTO_ACCEPT_In_A4        — auto-accept rows matching Annexure-4
  AUTO_ACCEPT_Not_In_A4    — auto-accept rows NOT in Annexure-4
  HUMAN_REVIEW_In_A4       — human-review rows matching Annexure-4
  HUMAN_REVIEW_Not_In_A4   — human-review rows NOT in Annexure-4

Extra columns on In_A4 sheets:
  ANNEXURE-4 MATCH, MATCH_TYPE (EXACT|SUFFIX_PREFIX|FUZZY), MATCH_FLAG (OK|DUPLICATE|SUSPICIOUS)
"""
from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def norm(t: str) -> str:
    """Normalize tag for matching (same rules as eval_coverage.py)."""
    s = (t or "").upper()
    for d in ("—", "–", "―", "−"):
        s = s.replace(d, "-")
    s = re.sub(r'[\s\-"“”‘’`\']+', "", s)
    s = s.replace("IN", "")
    return s


def load_annexure4_tags(path: str) -> tuple[set[str], dict[str, str]]:
    """Return normalized tag set and norm → original Annexure-4 tag map."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    norms: set[str] = set()
    norm_to_orig: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[2]:
            orig = str(row[2]).strip()
            n = norm(orig)
            norms.add(n)
            norm_to_orig.setdefault(n, orig)
    wb.close()
    return norms, norm_to_orig


def _is_digit_suffix_fragment(short: str, long: str) -> bool:
    """True when short is a prefix of long with only extra digits (e.g. VBV22 → VBV2246)."""
    if not long.startswith(short) or len(long) <= len(short):
        return False
    return long[len(short):].isdigit()


def annexure_match(tag: str, norm_to_orig: dict[str, str]) -> tuple[str, str, str]:
    """
    Return (annexure_tag, match_type, match_flag) for a pipeline tag.
    match_type: EXACT | SUFFIX_PREFIX | NONE
    match_flag: OK | SUSPICIOUS | empty (no match)
    """
    n = norm(tag)
    if not n:
        return "", "NONE", ""

    if n in norm_to_orig:
        return norm_to_orig[n], "EXACT", "OK"

    # Suffix match: pipeline tag missing area prefix (V-201 → V-V-201, TE-212 → V-TW-212)
    suffix_hits: list[tuple[str, str]] = []
    for an, orig in norm_to_orig.items():
        if len(an) > len(n) and an.endswith(n) and (len(an) - len(n)) <= 3:
            if not _is_digit_suffix_fragment(n, an):
                suffix_hits.append((orig, an))

    if len(suffix_hits) == 1:
        return suffix_hits[0][0], "SUFFIX_PREFIX", "OK"
    if len(suffix_hits) > 1:
        # Ambiguous — e.g. V-201 could hit multiple; pick closest length
        suffix_hits.sort(key=lambda x: len(norm(x[0])))
        return suffix_hits[0][0], "SUFFIX_PREFIX", "SUSPICIOUS"

    return "", "NONE", ""


def read_sheet_rows(path: str, sheet_name: str) -> tuple[list, list]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    return list(rows[0]), rows[1:]


def write_comparison_sheet(ws, headers: list, rows: list[list]):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(
            max(len(str(headers[col - 1])) + 2, 12), 40
        )
    ws.freeze_panes = "A2"


def split_rows(
    headers: list,
    data_rows: list,
    tag_col_idx: int,
    norm_to_orig: dict[str, str],
    sheet_label: str,
) -> tuple[list[list], list[list], dict]:
    """
    Split rows into in-A4 / not-in-A4.
    Returns (in_rows, not_rows, stats_dict).
    """
    extra = ["ANNEXURE-4 MATCH", "MATCH_TYPE", "MATCH_FLAG"]
    in_a4: list[list] = []
    not_in_a4: list[list] = []
    annexure_hits: dict[str, list[str]] = defaultdict(list)  # annexure → [pipeline tags]

    for i, row in enumerate(data_rows, start=1):
        if not row or not row[tag_col_idx]:
            continue
        tag = str(row[tag_col_idx]).strip()
        matched, mtype, mflag = annexure_match(tag, norm_to_orig)

        if matched:
            annexure_hits[matched].append(tag)
            # Mark duplicates after first occurrence
            if len(annexure_hits[matched]) > 1:
                mflag = "DUPLICATE"
            out = [i, sheet_label] + list(row) + [matched, mtype, mflag]
            in_a4.append(out)
        else:
            out = [i, sheet_label] + list(row)
            not_in_a4.append(out)

    stats = {
        "rows_in": len(in_a4),
        "rows_not": len(not_in_a4),
        "unique_annexure": len(annexure_hits),
        "annexure_hits": dict(annexure_hits),
    }
    return in_a4, not_in_a4, stats


def build_summary_rows(
    register_count: int,
    norm_to_orig: dict[str, str],
    aa_stats: dict,
    hr_stats: dict,
) -> list[list]:
    """Build SUMMARY sheet rows."""
    all_hits: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for stats, label in ((aa_stats, "AUTO_ACCEPT"), (hr_stats, "HUMAN_REVIEW")):
        for annex, tags in stats["annexure_hits"].items():
            for t in tags:
                all_hits[annex].append((label, t))

    covered = set(all_hits.keys())
    missing = sorted(set(norm_to_orig.values()) - covered)

    rows = [
        ["Metric", "Value"],
        ["Annexure-4 register tags", register_count],
        ["Unique register tags in final_tags (AA+HR)", len(covered)],
        ["Register tags NOT in final_tags", len(missing)],
        ["", ""],
        ["AUTO_ACCEPT rows matching A4", aa_stats["rows_in"]],
        ["AUTO_ACCEPT unique register tags", aa_stats["unique_annexure"]],
        ["AUTO_ACCEPT rows NOT in A4", aa_stats["rows_not"]],
        ["HUMAN_REVIEW rows matching A4", hr_stats["rows_in"]],
        ["HUMAN_REVIEW unique register tags", hr_stats["unique_annexure"]],
        ["HUMAN_REVIEW rows NOT in A4", hr_stats["rows_not"]],
        ["", ""],
        ["Duplicate register matches (AA)", sum(1 for v in aa_stats["annexure_hits"].values() if len(v) > 1)],
        ["Suspicious matches (AA+HR)", "see MATCH_FLAG=SUSPICIOUS in In_A4 sheets"],
        ["", ""],
        ["MISSING FROM final_tags", ""],
    ]
    for m in missing:
        rows.append([m, "not found in AUTO_ACCEPT or HUMAN_REVIEW"])

    rows.append(["", ""])
    rows.append(["DUPLICATE REGISTER MATCHES (same A4 tag, multiple rows)", ""])
    for annex in sorted(all_hits):
        if len(all_hits[annex]) > 1:
            detail = "; ".join(f"{lbl}:{t}" for lbl, t in all_hits[annex])
            rows.append([annex, detail])

    return rows


def main():
    ap = argparse.ArgumentParser(description="Compare final_tags.xlsx vs Annexure-4")
    ap.add_argument("--final", default="output/final_tags.xlsx")
    ap.add_argument("--register", default="ANNEXURE-4_4224-MGDV-6-50-2004-001-C.xlsx")
    ap.add_argument("--out", default="output/final_tags_vs_annexure4.xlsx")
    args = ap.parse_args()

    annexure_norms, norm_to_orig = load_annexure4_tags(args.register)

    aa_headers, aa_rows = read_sheet_rows(args.final, "AUTO_ACCEPT")
    hr_headers, hr_rows = read_sheet_rows(args.final, "HUMAN_REVIEW")
    tag_idx = aa_headers.index("TAG NUMBER")

    extra_cols = ["ANNEXURE-4 MATCH", "MATCH_TYPE", "MATCH_FLAG"]
    in_hdr = ["ROW#", "SHEET"] + list(aa_headers) + extra_cols

    not_hdr = ["ROW#", "SHEET"] + list(aa_headers)

    aa_in, aa_not, aa_stats = split_rows(aa_headers, aa_rows, tag_idx, norm_to_orig, "AUTO_ACCEPT")
    hr_in, hr_not, hr_stats = split_rows(hr_headers, hr_rows, tag_idx, norm_to_orig, "HUMAN_REVIEW")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary_rows = build_summary_rows(len(annexure_norms), norm_to_orig, aa_stats, hr_stats)
    ws_sum = wb.create_sheet("SUMMARY")
    write_comparison_sheet(ws_sum, summary_rows[0], summary_rows[1:])

    sheets = [
        ("AUTO_ACCEPT_In_A4", in_hdr, aa_in),
        ("AUTO_ACCEPT_Not_In_A4", not_hdr, aa_not),
        ("HUMAN_REVIEW_In_A4", in_hdr, hr_in),
        ("HUMAN_REVIEW_Not_In_A4", not_hdr, hr_not),
    ]
    for name, hdrs, rows in sheets:
        ws = wb.create_sheet(name)
        write_comparison_sheet(ws, hdrs, rows)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    unique_covered = len(set(aa_stats["annexure_hits"]) | set(hr_stats["annexure_hits"]))
    missing_n = len(annexure_norms) - unique_covered

    print(f"\n=== Annexure-4 comparison → {out_path} ===")
    print(f"  Annexure-4 register tags          : {len(annexure_norms)}")
    print(f"  Unique register tags in output    : {unique_covered}  ({unique_covered}/{len(annexure_norms)} covered)")
    print(f"  Register tags missing from output : {missing_n}")
    print()
    print("  AUTO_ACCEPT rows in A4            : {rows}  (unique register: {uniq})".format(
        rows=aa_stats["rows_in"], uniq=aa_stats["unique_annexure"]))
    print("  AUTO_ACCEPT rows not in A4        : {rows}".format(rows=aa_stats["rows_not"]))
    print("  HUMAN_REVIEW rows in A4           : {rows}  (unique register: {uniq})".format(
        rows=hr_stats["rows_in"], uniq=hr_stats["unique_annexure"]))
    print("  HUMAN_REVIEW rows not in A4       : {rows}".format(rows=hr_stats["rows_not"]))
    print()
    dup_aa = sum(1 for v in aa_stats["annexure_hits"].values() if len(v) > 1)
    if dup_aa:
        print(f"  ⚠ Duplicate A4 matches in AUTO_ACCEPT : {dup_aa} register tags have >1 row")
    susp = sum(1 for r in aa_in + hr_in if r[-1] == "SUSPICIOUS")
    if susp:
        print(f"  ⚠ Suspicious fuzzy matches              : {susp} (see MATCH_FLAG column)")
    print(f"\n  Open sheet SUMMARY for missing tags and duplicate detail.")


if __name__ == "__main__":
    main()
