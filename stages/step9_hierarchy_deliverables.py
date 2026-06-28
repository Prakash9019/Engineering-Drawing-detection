#!/usr/bin/env python3
"""
step9_hierarchy_deliverables.py — Engineer-friendly hierarchy deliverables
===========================================================================

Turns the machine-readable `step5b2_hierarchy_full.json` into auditable,
presentation-ready artefacts for CDCI engineering review.

INPUT
    --hierarchy  output/step5b2_hierarchy_full.json   (falls back to step5b2_hierarchy.json)
    --context    output/drawing_context.json           (Plant/Area/drawing metadata)
    --out        output/

OUTPUTS
    final_hierarchy.xlsx              Deliverable 1 — 6 sheets (Equipment Hierarchy,
                                      Parent-Child, Functional Location, Cross-Drawing,
                                      Orphan Nodes, Statistics)
    hierarchy_viewer.html             Deliverable 2 + 3 — interactive tree + search +
                                      node detail + Relationship Explorer tab
    hierarchy_graph.html              Deliverable 4 — self-contained colour-coded
                                      force-directed graph (no external/CDN deps)
    hierarchy_validation_report.xlsx  Deliverable 5 — validation rule checks

DATA-PROVENANCE NOTES (honest mapping — read before trusting columns)
    The source P&ID has NO explicit Plant/Area/Unit/Functional-Location codes, so:
      • Plant / Area / Unit are DERIVED from the drawing number
        (e.g. 4224-MGDV-6-50-2004 → Plant=MGDV, Area=6-50, Unit=sheet).
      • System      = the root equipment of the node's hierarchy tree (root_system).
      • Functional Location = dotted path Plant.Area.<ancestor tags>.<tag> (synthesised).
    Parent→Child backbone comes from step5b2 `direct_parent`/`children`.
    Relationship types are mapped from graph edge `rel`:
        MOUNTED_ON→support_function, CONTAINED_WITHIN→contains, SIGNAL_TO→controls,
        MONITORS→monitors, CONNECTED_TO→connected_to; hierarchy backbone→part_of;
        directed flow→feeds; control-loop roles→controls/monitors.
    Pipelines / junctions are "virtual" connector nodes (no physical tag).
    Cross-drawing references are detected heuristically from tags/contexts that look
    like other drawing numbers; if none are present the sheet is empty (expected on
    a single-sheet extraction).

This script is READ-ONLY over pipeline outputs — it never re-runs extraction.
"""

import argparse
import json
import math
import re
from collections import Counter, defaultdict, deque
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Relationship-type mapping (graph edge rel → business relationship)
# ─────────────────────────────────────────────────────────────────────────────
REL_TYPE = {
    "MOUNTED_ON":       "support_function",
    "CONTAINED_WITHIN": "contains",
    "SIGNAL_TO":        "controls",
    "MONITORS":         "monitors",
    "CONNECTED_TO":     "connected_to",
}
# rels intentionally excluded from the business parent-child sheet (spatial/structural noise)
NOISE_RELS = {"JUNCTION_OF", "ADJACENT_TO"}

VIRTUAL_KINDS = {"pipeline", "junction"}
CONTROL_KINDS = {"valve"}      # valves / actuators = control devices (orange)


# ─────────────────────────────────────────────────────────────────────────────
# Load + index
# ─────────────────────────────────────────────────────────────────────────────
def resolve_hierarchy_path(path_arg, out_dir):
    if path_arg:
        return path_arg
    full = Path(out_dir) / "step5b2_hierarchy_full.json"
    plain = Path(out_dir) / "step5b2_hierarchy.json"
    if full.exists():
        return str(full)
    if plain.exists():
        print("WARNING: step5b2_hierarchy_full.json not found — using step5b2_hierarchy.json")
        return str(plain)
    return str(full)


def derive_location(ctx):
    """Derive Plant / Area / Unit from the drawing number (documented as synthetic)."""
    dn = (ctx or {}).get("drawing_number", "") or ""
    parts = dn.split("-")
    plant = parts[1] if len(parts) > 1 else "PLANT"
    area = "-".join(parts[2:4]) if len(parts) >= 4 else (parts[0] if parts else "AREA")
    unit = (ctx or {}).get("sheet_number", "") or (parts[-1] if parts else "")
    return plant, area, unit


class HierModel:
    def __init__(self, H, ctx):
        self.H = H
        self.ctx = ctx or {}
        self.plant, self.area, self.unit = derive_location(ctx)
        self.drawing = self.ctx.get("drawing_number", "")
        self.sheet = self.ctx.get("sheet_number", "")
        self.rev = self.ctx.get("revision_code", "")
        self.source_drawing = f"{self.drawing} Sh.{self.sheet} Rev.{self.rev}".strip()

        # hierarchy records (tagged candidates only)
        self.hier = {h["node_id"]: h for h in H.get("hierarchy", [])}
        # enriched candidate detail (symbol_name, functional_context, confidence)
        self.ec = {e["candidate_id"]: e for e in H.get("enriched_candidates", [])}
        # full graph node index (incl. pipelines/junctions)
        self.gnodes = {n["node_id"]: n for n in H.get("graph", {}).get("nodes", [])}
        self.edges = H.get("graph", {}).get("edges", [])
        self.control_loops = H.get("control_loops", [])

        # control-loop membership: node_id → loop_id
        self.node_loop = {}
        self.loop_roles = {}
        for cl in self.control_loops:
            for nid in cl.get("member_ids", []):
                self.node_loop[nid] = cl["loop_id"]
            self.loop_roles[cl["loop_id"]] = cl.get("roles", {})

        self._eff_cache = {}

    def effective_parent(self, nid):
        """PURE READER — parent comes straight from step5b2's precomputed fields.

        FIX 3 (Part B): all hierarchy-repair logic (MOUNTED_ON resolution and
        pipeline line-collapse) now lives in step5b2, which writes the resolved
        EQUIPMENT parent into each hierarchy record as `equipment_parent_id`.
        Step9 no longer recomputes anything — it just reads, so the JSON, the
        Excel and the graph cannot disagree.

        Returns (parent_id_or_None, None). `via_line` is retired (always None);
        the 2-tuple shape is kept so existing call sites still unpack cleanly.
          1. direct_parent that is itself a real tagged component → use it.
          2. else the precomputed equipment_parent_id (step5b2 already collapsed
             pipelines / resolved MOUNTED_ON to a single equipment).
          3. else (None, None) = truly unparented."""
        if nid in self._eff_cache:
            return self._eff_cache[nid]
        h = self.hier.get(nid, {})
        res = (None, None)
        dp = h.get("direct_parent")
        if dp and dp in self.hier:
            res = (dp, None)
        else:
            ep = h.get("equipment_parent_id")
            if ep and ep in self.hier:
                res = (ep, None)
        self._eff_cache[nid] = res
        return res

    # ── label / attribute helpers ──
    def label(self, nid):
        if nid in self.hier:
            return self.hier[nid].get("tag_text") or nid
        g = self.gnodes.get(nid)
        if g:
            return g.get("tag_text") or g.get("ref") or nid
        return nid

    def kind(self, nid):
        if nid in self.hier:
            return self.hier[nid].get("kind", "")
        g = self.gnodes.get(nid)
        return g.get("kind", "") if g else ""

    def symbol_name(self, nid):
        e = self.ec.get(nid, {})
        return e.get("symbol_name", "") or ""

    def description(self, nid):
        e = self.ec.get(nid, {})
        return (e.get("functional_context") or e.get("symbol_name") or "").strip()

    def confidence(self, nid):
        e = self.ec.get(nid, {})
        for k in ("vision_confidence", "association_confidence", "ocr_confidence"):
            v = e.get(k)
            if v:
                return round(float(v), 2)
        return ""

    def ancestor_ids(self, nid):
        """Walk EFFECTIVE parents up to the root; return [root..parent] id list (cycle-safe)."""
        chain, seen, cur = [], set(), self.effective_parent(nid)[0]
        while cur and cur not in seen:
            seen.add(cur)
            chain.append(cur)
            cur = self.effective_parent(cur)[0]
        chain.reverse()
        return chain

    def ancestor_tags(self, nid):
        """Ancestor tag labels, collapsing consecutive duplicates (duplicate-node artefacts)."""
        out = []
        for i in self.ancestor_ids(nid):
            lbl = self.label(i)
            if not out or out[-1] != lbl:
                out.append(lbl)
        return out

    def effective_children(self, nid):
        if not hasattr(self, "_eff_children"):
            self._eff_children = defaultdict(list)
            for cid in self.hier:
                p = self.effective_parent(cid)[0]
                if p:
                    self._eff_children[p].append(cid)
        return self._eff_children.get(nid, [])

    def system_of(self, nid):
        """Nearest equipment in the effective-ancestor chain (root preferred); else self/UNASSIGNED."""
        chain = self.ancestor_ids(nid)
        for i in chain:                      # root → parent order; first equipment = system
            if self.hier.get(i, {}).get("kind") == "equipment":
                return self.label(i)
        if self.hier.get(nid, {}).get("kind") == "equipment":
            return self.label(nid)
        return chain and self.label(chain[0]) or "UNASSIGNED"

    def functional_location(self, nid):
        chain = self.ancestor_tags(nid) + [self.label(nid)]
        return f"{self.plant}.{self.area}." + ".".join(chain)

    def is_virtual(self, nid):
        return self.kind(nid) in VIRTUAL_KINDS


# ─────────────────────────────────────────────────────────────────────────────
# Relationship extraction
# ─────────────────────────────────────────────────────────────────────────────
def build_relationships(M: HierModel):
    """Return list of {parent, child, rel, source, confidence} business relationships."""
    rels = []
    seen = set()

    def add(parent_id, child_id, rtype, source, conf):
        key = (parent_id, child_id, rtype)
        if key in seen:
            return
        seen.add(key)
        rels.append({
            "parent_id": parent_id, "child_id": child_id,
            "parent": M.label(parent_id), "child": M.label(child_id),
            "rel": rtype, "source": source, "confidence": conf,
        })

    # 1. hierarchy backbone (effective parent → child = part_of; pipelines collapsed)
    for nid in M.hier:
        p, via = M.effective_parent(nid)
        if p:
            add(p, nid, "part_of",
                f"via_line:{M.label(via)}" if via else "hierarchy_tree", M.confidence(nid))

    # 2. graph edges — ONLY real tag↔tag relationships (skip virtual pipeline/junction
    #    endpoints; line connectivity is already in the part_of backbone via via_line).
    for e in M.edges:
        rel = e.get("rel")
        if rel in NOISE_RELS or rel not in REL_TYPE:
            continue
        f, t = e.get("from"), e.get("to")
        if f not in M.hier or t not in M.hier:
            continue
        rtype = REL_TYPE[rel]
        conf = e.get("confidence", "")
        src = e.get("evidence") or e.get("category") or rel
        if rel == "CONTAINED_WITHIN":
            # "f contained_within t"  → parent=t contains child=f
            add(t, f, "contains", src, conf)
        elif rel == "MOUNTED_ON":
            # "f mounted_on t" (instrument on equipment) → parent=t support_function child=f
            add(t, f, "support_function", src, conf)
        else:
            add(f, t, rtype, src, conf)

    # 3. control-loop roles → controls / monitors
    for cl in M.control_loops:
        loop_id = cl["loop_id"]
        roles = cl.get("roles", {})
        controllers = [m for m, r in roles.items() if "control" in r.lower()]
        # map tag→one node id present in this loop
        tag_to_id = {}
        for nid in cl.get("member_ids", []):
            tag_to_id[M.label(nid)] = nid
        for ctrl_tag in controllers:
            cid = tag_to_id.get(ctrl_tag)
            if not cid:
                continue
            for m_tag, role in roles.items():
                mid = tag_to_id.get(m_tag)
                if not mid or mid == cid:
                    continue
                rt = "monitors" if any(k in role.lower()
                                       for k in ("transmit", "element", "sens")) else "controls"
                add(cid, mid, rt, f"control_loop:{loop_id}", "")
    return rels


def detect_cross_drawing(M: HierModel):
    """Heuristic cross-drawing references: tags/contexts that look like other drawing numbers."""
    rows = []
    dn = M.drawing
    # a drawing-number-like token: digits-letters-… with >=3 dash groups, not this drawing
    pat = re.compile(r"\b(\d{3,4}-[A-Z]{2,5}-\d+-\d+(?:-\d+)?)\b")
    for nid, h in M.hier.items():
        tag = h.get("tag_text", "") or ""
        ec = M.ec.get(nid, {})
        ctx = ec.get("functional_context", "") or ""
        for text, where in ((tag, "tag"), (ctx, "context")):
            for m in pat.findall(text):
                if dn and m.startswith(dn.split("-")[0]) and m != dn:
                    rows.append({
                        "source_drawing": M.source_drawing, "source_tag": tag,
                        "ref_drawing": m, "ref_tag": "", "relationship": f"reference_in_{where}",
                    })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Validation rules (Deliverable 5)
# ─────────────────────────────────────────────────────────────────────────────
def validate(M: HierModel, rels):
    issues = []

    def add(itype, node, sev, reason, fix):
        issues.append({"issue": itype, "node": node, "severity": sev,
                       "reason": reason, "fix": fix})

    # parent counts from CONTAINS/PART_OF backbone for "multiple parents"
    parents_of = defaultdict(set)
    for r in rels:
        if r["rel"] in ("part_of", "contains"):
            parents_of[r["child_id"]].add(r["parent_id"])

    # multiple parents
    for cid, ps in parents_of.items():
        if len(ps) > 1:
            add("Multiple Parents", M.label(cid), "HIGH",
                f"{len(ps)} parents: {', '.join(sorted(M.label(p) for p in ps))}",
                "Confirm the true physical parent; remove spurious containment.")

    # cyclic hierarchy (effective-parent chain)
    for nid in M.hier:
        seen, cur = set(), nid
        while cur and cur not in seen:
            seen.add(cur)
            cur = M.effective_parent(cur)[0]
        if cur and cur in seen:
            add("Cyclic Hierarchy", M.label(nid), "HIGH",
                "effective-parent chain forms a cycle",
                "Break reciprocal CONTAINED_WITHIN; pick one direction.")

    # missing parent / system / functional location / isolation
    for nid, h in M.hier.items():
        tag = M.label(nid)
        pid, via = M.effective_parent(nid)
        truly_unparented = (not pid and not via and h.get("kind") != "equipment")
        if truly_unparented:
            add("Missing Parent", tag, "MEDIUM",
                "no equipment parent and not attached to any process line",
                "Manually assign parent equipment or system.")
        if h.get("is_isolated"):
            add("Disconnected Subgraph", tag, "MEDIUM",
                "isolated detection — no pipe/equipment/signal edge",
                "Verify detection; connect or reject as false positive.")
        # missing system only matters when the node is otherwise disconnected
        if truly_unparented and M.system_of(nid) == "UNASSIGNED":
            add("Missing System", tag, "LOW", "no process system resolved",
                "Assign to a process system.")

    # duplicate nodes (same tag, multiple node ids)
    by_tag = defaultdict(list)
    for nid, h in M.hier.items():
        by_tag[(h.get("tag_text") or "").upper()].append(nid)
    for tag, ids in by_tag.items():
        if tag and len(ids) > 1:
            add("Duplicate Nodes", tag, "MEDIUM",
                f"{len(ids)} nodes share tag '{tag}'",
                "Merge duplicate detections (step5d dedup).")

    # broken references (edge endpoint missing)
    node_ids = set(M.gnodes) | set(M.hier)
    for e in M.edges:
        for end in (e.get("from"), e.get("to")):
            if end not in node_ids:
                add("Broken Relationship", end, "HIGH",
                    f"edge {e.get('edge_id')} references unknown node",
                    "Re-run step5b2 on current extraction.")

    # invalid hierarchy depth (effective)
    for nid in M.hier:
        d = len(M.ancestor_ids(nid))
        if d > 12:
            add("Invalid Hierarchy Depth", M.label(nid), "LOW",
                f"effective depth={d}", "Investigate over-deep containment chain.")
    return issues


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SEV_FILL = {"HIGH": PatternFill("solid", fgColor="F8CBAD"),
            "MEDIUM": PatternFill("solid", fgColor="FFE699"),
            "LOW": PatternFill("solid", fgColor="E2EFDA")}


def write_sheet(ws, headers, rows, sev_col=None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r in rows:
        ws.append(r)
    # styling + width
    for ci, h in enumerate(headers, 1):
        width = max(len(str(h)), *(len(str(row[ci-1])) for row in rows)) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(width + 2, 10), 55)
    for ri in range(2, len(rows) + 2):
        for ci in range(1, len(headers) + 1):
            ws.cell(row=ri, column=ci).border = BORDER
            ws.cell(row=ri, column=ci).alignment = Alignment(vertical="top", wrap_text=True)
        if sev_col is not None:
            sev = ws.cell(row=ri, column=sev_col).value
            if sev in SEV_FILL:
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=ri, column=ci).fill = SEV_FILL[sev]
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# Deliverable 1 — final_hierarchy.xlsx
# ─────────────────────────────────────────────────────────────────────────────
def build_hierarchy_excel(M: HierModel, rels, cross_rows, out_path):
    wb = openpyxl.Workbook()

    # Sheet 1 — Equipment Hierarchy
    ws = wb.active
    ws.title = "Equipment Hierarchy"
    rows = []
    for nid, h in sorted(M.hier.items(), key=lambda kv: (len(M.ancestor_ids(kv[0])), M.label(kv[0]))):
        anc = M.ancestor_tags(nid)
        pid, via = M.effective_parent(nid)
        parent_disp = M.label(pid) if pid else (f"LINE:{M.label(via)}" if via else "NULL")
        rows.append([
            M.plant, M.area, M.system_of(nid), M.functional_location(nid),
            parent_disp, M.label(nid), M.description(nid), M.symbol_name(nid),
            len(anc), " > ".join(anc + [M.label(nid)]),
            M.source_drawing, M.confidence(nid),
        ])
    write_sheet(ws, ["Plant", "Area", "System", "Functional Location", "Parent Equipment",
                     "Equipment", "Equipment Description", "Equipment Type",
                     "Hierarchy Level", "Path", "Source Drawing", "Confidence"], rows)

    # Sheet 2 — Parent Child Relationships
    ws2 = wb.create_sheet("Parent Child Relationships")
    rrows = [[r["parent"], r["child"], r["rel"], r["source"], r["confidence"]] for r in rels]
    write_sheet(ws2, ["Parent Node", "Child Node", "Relationship Type", "Source", "Confidence"], rrows)

    # Sheet 3 — Functional Location Hierarchy
    ws3 = wb.create_sheet("Functional Location")
    frows = []
    for nid, h in sorted(M.hier.items(), key=lambda kv: M.label(kv[0])):
        anc = M.ancestor_tags(nid)
        system = anc[0] if anc else M.label(nid)
        sub = anc[1] if len(anc) > 1 else ""
        frows.append([M.plant, M.area, M.unit, system, sub,
                      M.label(nid), M.functional_location(nid)])
    write_sheet(ws3, ["Plant", "Area", "Unit", "System", "Sub-System",
                      "Equipment", "Functional Location"], frows)

    # Sheet 4 — Cross Drawing References
    ws4 = wb.create_sheet("Cross Drawing References")
    crows = [[r["source_drawing"], r["source_tag"], r["ref_drawing"], r["ref_tag"], r["relationship"]]
             for r in cross_rows]
    if not crows:
        crows = [["—", "—", "—", "—", "No cross-drawing references detected on this sheet"]]
    write_sheet(ws4, ["Source Drawing", "Source Tag", "Referenced Drawing",
                      "Referenced Tag", "Relationship"], crows)

    # Sheet 5 — Orphan Nodes
    ws5 = wb.create_sheet("Orphan Nodes")
    orows = []
    for nid, h in M.hier.items():
        pid, via = M.effective_parent(nid)
        # connected if it has an equipment parent OR is attached to a process line
        connected = bool(pid) or bool(via) or h.get("kind") == "equipment"
        isolated = bool(h.get("is_isolated"))
        no_parent = not pid and h.get("kind") != "equipment"   # no single equipment parent
        no_fl = M.functional_location(nid).endswith(".")
        no_sys = M.system_of(nid) == "UNASSIGNED"
        # an ORPHAN = truly disconnected: no equipment parent, no line, no edges
        orphan = isolated or (not connected)
        if orphan:
            reasons = []
            if not pid and not via: reasons.append("no parent / no line")
            if no_sys:    reasons.append("no system")
            if no_fl:     reasons.append("no functional location")
            if isolated:  reasons.append("isolated (no graph edge)")
            orows.append([M.label(nid), M.hier[nid].get("kind", ""),
                          "YES" if no_parent else "", "YES" if no_fl else "",
                          "YES" if no_sys else "", "YES" if isolated else "",
                          "; ".join(reasons) or "line-rooted (review)"])
    if not orows:
        orows = [["—", "—", "", "", "", "", "No orphan nodes"]]
    write_sheet(ws5, ["Node", "Kind", "No Parent Found", "No Functional Location",
                      "No System", "Broken Relationship", "Reason"], orows)

    # Sheet 6 — Hierarchy Statistics
    ws6 = wb.create_sheet("Hierarchy Statistics")
    stats = M.H.get("stats", {})
    kinds = Counter(h.get("kind") for h in M.hier.values())
    depth_max = max((len(M.ancestor_ids(nid)) for nid in M.hier), default=0)
    n_virtual = sum(1 for n in M.gnodes.values() if n.get("kind") in VIRTUAL_KINDS)
    n_orphan = len(orows) if orows and orows[0][0] != "—" else 0
    srows = [
        ["Total Nodes (tagged)", len(M.hier)],
        ["Total Graph Nodes (incl. virtual)", len(M.gnodes)],
        ["Total Relationships (business)", len(rels)],
        ["Total Graph Edges", len(M.edges)],
        ["Equipment Count", kinds.get("equipment", 0)],
        ["Instrument Count", kinds.get("instrument", 0)],
        ["Valve Count", kinds.get("valve", 0)],
        ["Piping Count", kinds.get("piping", 0)],
        ["Control Loops", len(M.control_loops)],
        ["Cross Drawing Links", len(cross_rows)],
        ["Orphan Nodes", n_orphan],
        ["Virtual Nodes (pipelines+junctions)", n_virtual],
        ["Hierarchy Depth (max)", depth_max],
        ["Isolated Detections", stats.get("n_candidates_isolated", "")],
    ]
    write_sheet(ws6, ["Metric", "Value"], srows)
    ws6.column_dimensions["A"].width = 40

    wb.save(out_path)
    return {"orphans": n_orphan, "relationships": len(rels), "virtual": n_virtual,
            "depth": depth_max, "cross": len(cross_rows)}


# ─────────────────────────────────────────────────────────────────────────────
# Deliverable 5 — hierarchy_validation_report.xlsx
# ─────────────────────────────────────────────────────────────────────────────
def build_validation_excel(issues, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Issues"
    rows = [[i["issue"], i["node"], i["severity"], i["reason"], i["fix"]] for i in issues]
    if not rows:
        rows = [["—", "—", "—", "No validation issues found", ""]]
    write_sheet(ws, ["Issue Type", "Node", "Severity", "Reason", "Suggested Fix"],
                rows, sev_col=3)

    ws2 = wb.create_sheet("Summary")
    by_type = Counter(i["issue"] for i in issues)
    by_sev = Counter(i["severity"] for i in issues)
    srows = [["TOTAL ISSUES", len(issues)]]
    srows += [[f"Severity: {k}", v] for k, v in sorted(by_sev.items())]
    srows += [[f"Type: {k}", v] for k, v in sorted(by_type.items())]
    write_sheet(ws2, ["Metric", "Count"], srows)
    ws2.column_dimensions["A"].width = 32
    wb.save(out_path)
    return by_sev


# ─────────────────────────────────────────────────────────────────────────────
# Build the JSON payload shared by both HTML viewers
# ─────────────────────────────────────────────────────────────────────────────
def build_view_payload(M: HierModel, rels, issues):
    nodes = []
    for nid, h in M.hier.items():
        nodes.append({
            "id": nid,
            "tag": M.label(nid),
            "kind": h.get("kind", ""),
            "type": M.symbol_name(nid),
            "desc": M.description(nid),
            "parent": M.effective_parent(nid)[0],
            "depth": len(M.ancestor_ids(nid)),
            "system": M.system_of(nid),
            "fl": M.functional_location(nid),
            "conf": M.confidence(nid),
            "isolated": bool(h.get("is_isolated")),
            "loop": M.node_loop.get(nid, ""),
            "children": M.effective_children(nid),
            "drawing": M.source_drawing,
        })
    # adjacency for relationship explorer
    rel_out = [{"p": r["parent_id"], "c": r["child_id"], "rel": r["rel"],
                "ptag": r["parent"], "ctag": r["child"], "src": r["source"]} for r in rels]
    return {
        "meta": {"plant": M.plant, "area": M.area, "unit": M.unit,
                 "drawing": M.source_drawing, "n_nodes": len(nodes),
                 "n_rels": len(rel_out), "n_loops": len(M.control_loops),
                 "n_issues": len(issues)},
        "nodes": nodes,
        "rels": rel_out,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Deliverable 2 + 3 — hierarchy_viewer.html (tree + search + relationship explorer)
# ─────────────────────────────────────────────────────────────────────────────
def build_viewer_html(payload, out_path):
    data_json = json.dumps(payload, separators=(",", ":"))
    html = _VIEWER_TEMPLATE.replace("/*__DATA__*/", data_json)
    Path(out_path).write_text(html, encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# Deliverable 4 — hierarchy_graph.html (self-contained canvas force graph)
# ─────────────────────────────────────────────────────────────────────────────
def build_graph_html(payload, out_path):
    data_json = json.dumps(payload, separators=(",", ":"))
    html = _GRAPH_TEMPLATE.replace("/*__DATA__*/", data_json)
    Path(out_path).write_text(html, encoding="utf-8")


# ═════════════════════════════════════════════════════════════════════════════
# HTML templates (self-contained — no external CSS/JS/CDN)
# ═════════════════════════════════════════════════════════════════════════════
_VIEWER_TEMPLATE = r"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>CDCI Hierarchy Viewer</title>
<style>
*{box-sizing:border-box} body{margin:0;font:14px/1.4 -apple-system,Segoe UI,Roboto,Arial,sans-serif;color:#1b2733;background:#f4f6f9}
header{background:#1f4e78;color:#fff;padding:12px 18px;display:flex;align-items:center;gap:18px;flex-wrap:wrap}
header h1{font-size:17px;margin:0} header .meta{font-size:12px;opacity:.85}
.tabs{display:flex;gap:4px;background:#163a5a;padding:0 12px}
.tabs button{background:none;border:0;color:#cfe0f0;padding:10px 16px;cursor:pointer;font-size:13px;border-bottom:3px solid transparent}
.tabs button.active{color:#fff;border-bottom-color:#7fb0e0;font-weight:600}
.wrap{display:flex;height:calc(100vh - 96px)}
.left{flex:1;overflow:auto;padding:12px 16px;background:#fff;border-right:1px solid #e0e5ec}
.right{width:360px;overflow:auto;padding:16px;background:#fafbfc}
.search{display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap}
.search input,.search select{padding:7px 9px;border:1px solid #c4ccd6;border-radius:6px;font-size:13px}
.search input{flex:1;min-width:160px}
ul.tree{list-style:none;margin:0;padding-left:16px}
ul.tree.root{padding-left:0}
li.node{margin:1px 0}
.row{display:flex;align-items:center;gap:6px;padding:3px 6px;border-radius:5px;cursor:pointer}
.row:hover{background:#eef4fb}
.row.sel{background:#d7e8fb;outline:1px solid #7fb0e0}
.row.hl-parent{background:#fff3cd} .row.hl-child{background:#e2f0d9}
.tog{width:14px;text-align:center;color:#7a8aa0;font-size:11px;user-select:none}
.tag{font-weight:600} .badge{font-size:10px;padding:1px 6px;border-radius:10px;color:#fff}
.b-equipment{background:#2f6fb0}.b-instrument{background:#3a9b56}.b-valve{background:#e08a2f}.b-piping{background:#c08a2f}
.iso{color:#c0392b;font-size:11px} .small{color:#7a8aa0;font-size:11px}
.detail h3{margin:.2em 0;font-size:15px} .detail .k{color:#7a8aa0;font-size:11px;text-transform:uppercase;letter-spacing:.04em;margin-top:10px}
.detail .v{font-size:14px;word-break:break-word} .chip{display:inline-block;background:#eef2f7;border:1px solid #d6deea;border-radius:6px;padding:2px 8px;margin:2px 3px 0 0;font-size:12px;cursor:pointer}
.chip:hover{background:#dceafa} .hidden{display:none}
.exp .box{background:#fff;border:1px solid #e0e5ec;border-radius:8px;padding:12px 14px;margin-bottom:12px}
.exp .lbl{color:#7a8aa0;font-size:11px;text-transform:uppercase;letter-spacing:.04em} .exp .val{font-size:15px;font-weight:600}
footer{font-size:11px;color:#7a8aa0;padding:6px 16px;background:#fff;border-top:1px solid #e0e5ec}
</style></head><body>
<header><h1>CDCI Hierarchy Viewer</h1><span class="meta" id="meta"></span></header>
<div class="tabs"><button id="tab-tree" class="active" onclick="showTab('tree')">Tree View</button>
<button id="tab-exp" onclick="showTab('exp')">Relationship Explorer</button></div>

<div id="view-tree" class="wrap">
  <div class="left">
    <div class="search">
      <input id="q" placeholder="Search tag / equipment / functional location…" oninput="filterTree()">
      <select id="qmode" onchange="filterTree()">
        <option value="tag">By Tag</option><option value="equip">By Equipment</option>
        <option value="fl">By Functional Location</option></select>
      <button onclick="expandAll(true)">Expand all</button><button onclick="expandAll(false)">Collapse all</button>
    </div>
    <ul id="tree" class="tree root"></ul>
  </div>
  <div class="right"><div id="detail" class="detail"><p class="small">Click a node to see details.</p></div></div>
</div>

<div id="view-exp" class="wrap hidden">
  <div class="left" style="max-width:420px">
    <div class="search"><input id="eq" placeholder="Type a tag to explore…" oninput="expSearch()"></div>
    <ul id="elist" class="tree root"></ul>
  </div>
  <div class="right exp" id="expbox"><p class="small">Select a tag on the left.</p></div>
</div>
<footer id="foot"></footer>

<script>
const DATA=/*__DATA__*/;
const byId={}, childrenMap={}; DATA.nodes.forEach(n=>{byId[n.id]=n;});
DATA.nodes.forEach(n=>{const p=n.parent; if(p&&byId[p]){(childrenMap[p]=childrenMap[p]||[]).push(n.id);}});
const roots=DATA.nodes.filter(n=>!n.parent||!byId[n.parent]).map(n=>n.id);
// relationship adjacency
const relsBy={}; DATA.rels.forEach(r=>{(relsBy[r.p]=relsBy[r.p]||[]).push({dir:'out',rel:r.rel,id:r.c,tag:r.ctag});
 (relsBy[r.c]=relsBy[r.c]||[]).push({dir:'in',rel:r.rel,id:r.p,tag:r.ptag});});
document.getElementById('meta').textContent=`${DATA.meta.drawing} · ${DATA.meta.n_nodes} nodes · ${DATA.meta.n_rels} relationships · ${DATA.meta.n_loops} control loops`;
document.getElementById('foot').textContent=`Plant ${DATA.meta.plant} · Area ${DATA.meta.area} · Unit ${DATA.meta.unit} · ${DATA.meta.n_issues} validation issues — generated from step5b2_hierarchy_full.json`;

function badge(k){return `<span class="badge b-${k}">${k}</span>`;}
function nodeRow(n){
  const kids=childrenMap[n.id]||[];
  const tog=kids.length?`<span class="tog" data-t>▶</span>`:`<span class="tog"></span>`;
  return `<li class="node" data-id="${n.id}"><div class="row" data-row onclick="pick('${n.id}',event)">
    ${tog}<span class="tag">${esc(n.tag)}</span>${badge(n.kind)}
    ${n.isolated?'<span class="iso">⚠ isolated</span>':''}
    <span class="small">${esc(n.type||'')}</span></div>
    ${kids.length?`<ul class="tree hidden"></ul>`:''}</li>`;
}
function esc(s){return (s==null?'':String(s)).replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));}

function buildTree(){
  const root=document.getElementById('tree'); root.innerHTML='';
  roots.sort((a,b)=>byId[a].tag.localeCompare(byId[b].tag)).forEach(id=>root.insertAdjacentHTML('beforeend',nodeRow(byId[id])));
  wire(root);
}
function wire(container){
  container.querySelectorAll(':scope > li.node').forEach(li=>{
    const tog=li.querySelector(':scope > .row > .tog[data-t]');
    if(tog){tog.onclick=(e)=>{e.stopPropagation();toggle(li);};}
  });
}
function toggle(li){
  const ul=li.querySelector(':scope > ul'); if(!ul)return;
  const tog=li.querySelector(':scope > .row > .tog');
  if(ul.classList.contains('hidden')){
    if(!ul.dataset.loaded){const id=li.dataset.id;(childrenMap[id]||[]).map(c=>byId[c])
      .sort((a,b)=>a.tag.localeCompare(b.tag)).forEach(c=>ul.insertAdjacentHTML('beforeend',nodeRow(c)));
      ul.dataset.loaded=1; wire(ul);}
    ul.classList.remove('hidden'); tog.textContent='▼';
  } else {ul.classList.add('hidden'); tog.textContent='▶';}
}
function expandAll(open){
  document.querySelectorAll('#tree li.node').forEach(li=>{
    const ul=li.querySelector(':scope > ul'); if(!ul)return;
    if(open && ul.classList.contains('hidden'))toggle(li);
    if(!open && !ul.classList.contains('hidden'))toggle(li);
  });
  if(open){let pass=0;const iv=setInterval(()=>{let did=false;
    document.querySelectorAll('#tree li.node > ul.hidden').forEach(ul=>{toggle(ul.parentElement);did=true;});
    if(!did||++pass>30)clearInterval(iv);},10);}
}
function pathTo(id){const p=[];let c=id;const seen=new Set();while(c&&byId[c]&&!seen.has(c)){seen.add(c);p.unshift(c);c=byId[c].parent;}return p;}
function pick(id,e){
  if(e)e.stopPropagation();
  document.querySelectorAll('#tree .row').forEach(r=>r.classList.remove('sel','hl-parent','hl-child'));
  // ensure path expanded
  const path=pathTo(id);
  path.forEach(pid=>{const li=document.querySelector(`#tree li.node[data-id="${pid}"]`);
    if(li){const ul=li.querySelector(':scope > ul'); if(ul&&ul.classList.contains('hidden'))toggle(li);}});
  const li=document.querySelector(`#tree li.node[data-id="${id}"]`);
  if(li){const row=li.querySelector(':scope > .row');row.classList.add('sel');row.scrollIntoView({block:'center'});}
  const n=byId[id];
  if(n.parent){const pr=document.querySelector(`#tree li.node[data-id="${n.parent}"] > .row`);if(pr)pr.classList.add('hl-parent');}
  (childrenMap[id]||[]).forEach(cid=>{const cr=document.querySelector(`#tree li.node[data-id="${cid}"] > .row`);if(cr)cr.classList.add('hl-child');});
  showDetail(n);
}
function showDetail(n){
  const kids=(childrenMap[n.id]||[]).map(c=>byId[c].tag);
  const rc=(relsBy[n.id]||[]).length;
  const path=pathTo(n.id).map(i=>esc(byId[i].tag)).join(' &rsaquo; ');
  document.getElementById('detail').innerHTML=`
   <h3>${esc(n.tag)} ${badge(n.kind)}</h3>
   <div class="k">Description</div><div class="v">${esc(n.desc||'—')}</div>
   <div class="k">Equipment Type</div><div class="v">${esc(n.type||'—')}</div>
   <div class="k">Parent</div><div class="v">${n.parent&&byId[n.parent]?`<span class="chip" onclick="pick('${n.parent}')">${esc(byId[n.parent].tag)}</span>`:'NULL (top-level)'}</div>
   <div class="k">Children (${kids.length})</div><div class="v">${kids.length?(childrenMap[n.id]||[]).map(c=>`<span class="chip" onclick="pick('${c}')">${esc(byId[c].tag)}</span>`).join(''):'None'}</div>
   <div class="k">Full Path</div><div class="v small">${path}</div>
   <div class="k">System</div><div class="v">${esc(n.system||'—')}</div>
   <div class="k">Functional Location</div><div class="v small">${esc(n.fl)}</div>
   <div class="k">Control Loop</div><div class="v">${esc(n.loop||'—')}</div>
   <div class="k">Drawing</div><div class="v">${esc(n.drawing)}</div>
   <div class="k">Confidence</div><div class="v">${esc(n.conf)}</div>
   <div class="k">Relationship Count</div><div class="v">${rc}${n.isolated?' &nbsp;<span class="iso">⚠ isolated detection</span>':''}</div>`;
}
function filterTree(){
  const q=document.getElementById('q').value.trim().toLowerCase();
  const mode=document.getElementById('qmode').value;
  if(!q){buildTree();return;}
  const match=DATA.nodes.filter(n=>{
    if(mode==='tag')return n.tag.toLowerCase().includes(q);
    if(mode==='equip')return (n.kind==='equipment')&&(n.tag.toLowerCase().includes(q)||(n.desc||'').toLowerCase().includes(q));
    return (n.fl||'').toLowerCase().includes(q);
  });
  const root=document.getElementById('tree');root.innerHTML='';
  if(!match.length){root.innerHTML='<li class="small">No matches.</li>';return;}
  match.slice(0,400).forEach(n=>{
    root.insertAdjacentHTML('beforeend',
     `<li class="node" data-id="${n.id}"><div class="row" onclick="pick('${n.id}',event)">
       <span class="tog"></span><span class="tag">${esc(n.tag)}</span>${badge(n.kind)}
       <span class="small">${esc(n.fl)}</span></div></li>`);
  });
}
// Relationship Explorer
function expSearch(){
  const q=document.getElementById('eq').value.trim().toLowerCase();
  const list=document.getElementById('elist');list.innerHTML='';
  DATA.nodes.filter(n=>n.tag.toLowerCase().includes(q)).slice(0,300).forEach(n=>{
    list.insertAdjacentHTML('beforeend',
     `<li class="node"><div class="row" onclick="explore('${n.id}')"><span class="tog"></span>
       <span class="tag">${esc(n.tag)}</span>${badge(n.kind)}</div></li>`);
  });
}
function explore(id){
  const n=byId[id];
  const parent=n.parent&&byId[n.parent]?byId[n.parent]:null;
  const grand=parent&&parent.parent&&byId[parent.parent]?byId[parent.parent]:null;
  const kids=(childrenMap[id]||[]).map(c=>byId[c].tag);
  const rels=relsBy[id]||[];
  const grp=t=>rels.filter(r=>r.rel===t).map(r=>r.tag);
  const controls=grp('controls'), monitors=grp('monitors'), conn=grp('connected_to'),
        feeds=grp('feeds'), support=grp('support_function'), contains=grp('contains');
  function box(lbl,val){return `<div class="box"><div class="lbl">${lbl}</div><div class="val">${val||'None'}</div></div>`;}
  function chips(arr){return arr.length?arr.map(t=>`<span class="chip" onclick="explore(idOf('${esc(t)}'))">${esc(t)}</span>`).join(''):'None';}
  document.getElementById('expbox').innerHTML=
    `<h3 style="margin-top:0">Selected: ${esc(n.tag)} ${badge(n.kind)}</h3>`+
    box('Description',esc(n.desc||'—'))+
    box('Parent',parent?chips([parent.tag]):'NULL (top-level)')+
    box('Grand Parent',grand?chips([grand.tag]):'None')+
    box('Children',chips(kids))+
    box('Controls',chips(controls))+
    box('Monitors',chips(monitors))+
    box('Connected To',chips(conn))+
    box('Feeds',chips(feeds))+
    box('Mounted / Support',chips(support))+
    box('Contains',chips(contains))+
    box('Control Loop',esc(n.loop||'—'))+
    box('Functional Location',`<span class="small">${esc(n.fl)}</span>`)+
    box('Referenced In',esc(n.drawing));
}
function idOf(tag){const n=DATA.nodes.find(x=>x.tag===tag);return n?n.id:'';}
function showTab(t){
  document.getElementById('view-tree').classList.toggle('hidden',t!=='tree');
  document.getElementById('view-exp').classList.toggle('hidden',t!=='exp');
  document.getElementById('tab-tree').classList.toggle('active',t==='tree');
  document.getElementById('tab-exp').classList.toggle('active',t==='exp');
  if(t==='exp'&&!document.getElementById('elist').children.length)expSearch();
}
buildTree();
</script></body></html>"""


_GRAPH_TEMPLATE = r"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>CDCI Hierarchy Graph</title>
<style>
*{box-sizing:border-box}body{margin:0;font:13px -apple-system,Segoe UI,Roboto,Arial,sans-serif;background:#0f1623;color:#dfe7f2;overflow:hidden}
header{position:fixed;top:0;left:0;right:0;background:#16243a;padding:8px 14px;display:flex;gap:14px;align-items:center;flex-wrap:wrap;z-index:5;border-bottom:1px solid #24344f}
header h1{font-size:15px;margin:0}.legend{display:flex;gap:12px;font-size:12px;flex-wrap:wrap}
.legend span{display:flex;align-items:center;gap:5px}.dot{width:11px;height:11px;border-radius:50%}
input{padding:5px 8px;border-radius:6px;border:1px solid #2c3e5c;background:#0f1c30;color:#dfe7f2}
#info{position:fixed;right:0;top:46px;bottom:0;width:280px;background:#0d1726ee;border-left:1px solid #24344f;padding:14px;overflow:auto;z-index:4}
#info .k{color:#7c93b5;font-size:11px;text-transform:uppercase;margin-top:9px}#info .v{font-size:14px;word-break:break-word}
canvas{display:block}.hint{font-size:11px;color:#7c93b5}
</style></head><body>
<header><h1>CDCI Hierarchy Graph</h1>
<div class="legend">
<span><i class="dot" style="background:#3a7bd5"></i>Equipment</span>
<span><i class="dot" style="background:#3a9b56"></i>Instrument</span>
<span><i class="dot" style="background:#e08a2f"></i>Control Device</span>
<span><i class="dot" style="background:#9b59b6"></i>Virtual</span>
<span><i class="dot" style="background:#c0392b"></i>Orphan</span></div>
<input id="q" placeholder="highlight tag…" oninput="hl()">
<span class="hint">drag = pan · wheel = zoom · click node = info</span></header>
<div id="info"><p class="hint">Click a node.</p></div>
<canvas id="cv"></canvas>
<script>
const DATA=/*__DATA__*/;
const COL={equipment:'#3a7bd5',instrument:'#3a9b56',valve:'#e08a2f',piping:'#e08a2f',virtual:'#9b59b6',orphan:'#c0392b'};
const RELCOL={part_of:'#56708f',contains:'#56708f',controls:'#e08a2f',monitors:'#3a9b56',connected_to:'#3a7bd5',feeds:'#5dade2',support_function:'#9b59b6'};
const idx={};DATA.nodes.forEach((n,i)=>{idx[n.id]=i;});
const N=DATA.nodes.map(n=>({...n,x:Math.cos(idx[n.id])*300+Math.random()*40,y:Math.sin(idx[n.id])*300+Math.random()*40,vx:0,vy:0}));
const E=DATA.rels.filter(r=>idx[r.p]!=null&&idx[r.c]!=null).map(r=>({s:idx[r.p],t:idx[r.c],rel:r.rel}));
function color(n){if(n.isolated)return COL.orphan;return COL[n.kind]||COL.virtual;}
function rad(n){return n.kind==='equipment'?9:(n.kind==='instrument'?4.5:5.5);}
// degree
const deg={};E.forEach(e=>{deg[e.s]=(deg[e.s]||0)+1;deg[e.t]=(deg[e.t]||0)+1;});
// force sim
function step(){
  const k=0.012,rep=2600;
  for(let i=0;i<N.length;i++){let fx=0,fy=0;const a=N[i];
    for(let j=0;j<N.length;j++){if(i===j)continue;const b=N[j];let dx=a.x-b.x,dy=a.y-b.y;let d2=dx*dx+dy*dy+0.01;
      if(d2<90000){const f=rep/d2;fx+=dx*f;fy+=dy*f;}}
    fx+=-a.x*0.002;fy+=-a.y*0.002; a.vx=(a.vx+fx)*0.82;a.vy=(a.vy+fy)*0.82;}
  E.forEach(e=>{const a=N[e.s],b=N[e.t];let dx=b.x-a.x,dy=b.y-a.y;const d=Math.sqrt(dx*dx+dy*dy)||1;
    const f=(d-70)*k;dx/=d;dy/=d;a.vx+=dx*f;a.vy+=dy*f;b.vx-=dx*f;b.vy-=dy*f;});
  for(const a of N){if(a===drag)continue;a.x+=a.vx;a.y+=a.vy;}
}
const cv=document.getElementById('cv'),cx=cv.getContext('2d');
let view={x:0,y:0,z:1},W,Hh;
function resize(){W=cv.width=innerWidth;Hh=cv.height=innerHeight;view.x=W/2;view.y=Hh/2;}
window.onresize=resize;resize();
let hlTag='';
function draw(){
  cx.clearRect(0,0,W,Hh);cx.save();cx.translate(view.x,view.y);cx.scale(view.z,view.z);
  cx.lineWidth=0.7;
  E.forEach(e=>{const a=N[e.s],b=N[e.t];cx.strokeStyle=(RELCOL[e.rel]||'#34465f')+'88';
    cx.beginPath();cx.moveTo(a.x,a.y);cx.lineTo(b.x,b.y);cx.stroke();});
  N.forEach(n=>{cx.beginPath();cx.fillStyle=color(n);
    const r=rad(n)*(hlTag&&n.tag.toLowerCase().includes(hlTag)?1.9:1);
    cx.arc(n.x,n.y,r,0,7);cx.fill();
    if(n.kind==='equipment'||(hlTag&&n.tag.toLowerCase().includes(hlTag))){
      cx.fillStyle='#dfe7f2';cx.font='9px sans-serif';cx.fillText(n.tag,n.x+r+2,n.y+3);}});
  cx.restore();
}
let drag=null,pan=null;
let frames=0;function loop(){if(frames++<600)step();draw();requestAnimationFrame(loop);}loop();
// interaction
function toWorld(mx,my){return{x:(mx-view.x)/view.z,y:(my-view.y)/view.z};}
cv.onmousedown=e=>{const w=toWorld(e.clientX,e.clientY);let hit=null,hd=1e9;
  N.forEach(n=>{const d=(n.x-w.x)**2+(n.y-w.y)**2;if(d<hd&&d<160){hd=d;hit=n;}});
  if(hit){drag=hit;info(hit);}else{pan={x:e.clientX-view.x,y:e.clientY-view.y};}};
cv.onmousemove=e=>{if(drag){const w=toWorld(e.clientX,e.clientY);drag.x=w.x;drag.y=w.y;drag.vx=drag.vy=0;frames=Math.min(frames,400);}
  else if(pan){view.x=e.clientX-pan.x;view.y=e.clientY-pan.y;}};
window.onmouseup=()=>{drag=null;pan=null;};
cv.onwheel=e=>{e.preventDefault();const f=e.deltaY<0?1.1:0.9;view.z=Math.max(0.15,Math.min(4,view.z*f));};
function hl(){hlTag=document.getElementById('q').value.trim().toLowerCase();}
function info(n){const out=DATA.rels.filter(r=>r.p===n.id),inn=DATA.rels.filter(r=>r.c===n.id);
  document.getElementById('info').innerHTML=`<h3 style="margin:.2em 0">${esc(n.tag)}</h3>
   <div class="k">Kind</div><div class="v">${esc(n.kind)}${n.isolated?' ⚠ orphan':''}</div>
   <div class="k">Description</div><div class="v">${esc(n.desc||'—')}</div>
   <div class="k">System</div><div class="v">${esc(n.system||'—')}</div>
   <div class="k">Functional Location</div><div class="v" style="font-size:11px">${esc(n.fl)}</div>
   <div class="k">Out relationships (${out.length})</div><div class="v" style="font-size:12px">${out.map(r=>esc(r.rel)+'→'+esc(r.ctag)).join('<br>')||'—'}</div>
   <div class="k">In relationships (${inn.length})</div><div class="v" style="font-size:12px">${inn.map(r=>esc(r.ptag)+'→'+esc(r.rel)).join('<br>')||'—'}</div>`;}
function esc(s){return(s==null?'':String(s)).replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));}
</script></body></html>"""


# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Step 9: hierarchy deliverables (Excel + HTML)")
    ap.add_argument("--hierarchy", help="step5b2_hierarchy_full.json (default: auto-resolve in --out)")
    ap.add_argument("--context", default="output/drawing_context.json")
    ap.add_argument("--out", default="output")
    args = ap.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    hpath = resolve_hierarchy_path(args.hierarchy, args.out)
    print(f"Hierarchy source : {hpath}")
    H = json.load(open(hpath))
    ctx = {}
    if Path(args.context).exists():
        ctx = json.load(open(args.context))

    M = HierModel(H, ctx)
    rels = build_relationships(M)
    cross = detect_cross_drawing(M)
    issues = validate(M, rels)

    # Deliverable 1
    p1 = out / "final_hierarchy.xlsx"
    s = build_hierarchy_excel(M, rels, cross, p1)
    # Deliverable 5
    p5 = out / "hierarchy_validation_report.xlsx"
    sev = build_validation_excel(issues, p5)
    # Deliverable 2 + 3
    payload = build_view_payload(M, rels, issues)
    p2 = out / "hierarchy_viewer.html"
    build_viewer_html(payload, p2)
    # Deliverable 4
    p4 = out / "hierarchy_graph.html"
    build_graph_html(payload, p4)

    print("\n=== Step 9 — Hierarchy Deliverables ===")
    print(f"  Tagged nodes        : {len(M.hier)}")
    print(f"  Relationships       : {s['relationships']}")
    print(f"  Virtual nodes       : {s['virtual']}")
    print(f"  Orphan nodes        : {s['orphans']}")
    print(f"  Hierarchy depth     : {s['depth']}")
    print(f"  Cross-drawing refs  : {s['cross']}")
    print(f"  Validation issues   : {len(issues)}  ({dict(sev)})")
    print("\n  Outputs:")
    for p in (p1, p2, p4, p5):
        print(f"    {p}")


if __name__ == "__main__":
    main()
