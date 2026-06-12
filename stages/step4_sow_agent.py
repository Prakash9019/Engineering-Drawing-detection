#!/usr/bin/env python3
"""
step4_sow_agent.py — SOW Symbol Scope Intelligence Agent
=========================================================
CDCI P&ID Pipeline — Step 4

What this does
--------------
Reads the project Excel file (ANNEXURE-2_CDC-SYMBOLS_USED_AND_NOT_USED.xlsx)
containing two sheets:

  Sheet 1 — SYMBOL-USE        → 100 symbols tagged ALLOW / IN_SCOPE
  Sheet 2 — SYMBOL DONT USE   → 32 symbols tagged BLOCK / OUT_OF_SCOPE

For every symbol in both sheets:
  • Extracts the embedded symbol image from the Excel cell
  • Reads the symbol name / description
  • Infers ISA category, tag prefix patterns, visual shape descriptors
  • Uses Gemini 2.5 Flash to describe each symbol image → stores visual features
  • Builds a persistent Symbol Scope Memory JSON

Downstream agents call classify_symbol(symbol_name, visual_description)
to get ALLOW / BLOCK / UNSPECIFIED + full reasoning.

Pipeline
--------
  Phase 1: Parse Excel → extract images + names → build raw symbol list
  Phase 2: Gemini vision → describe each symbol image → add visual_features
  Phase 3: Persist Symbol Scope Memory → sow_symbol_memory.json
  Phase 4: Runtime classifier API (classify_symbol, check_tag_scope)

Outputs
-------
  sow_symbol_memory.json      — Full symbol scope knowledge base
  sow_scope_summary.txt       — Human-readable summary
  drawing_context.json        — Updated with SOW memory path

Usage
-----
  # Build the memory (run once per project):
  python step4_sow_agent.py build --excel ANNEXURE-2.xlsx --out output/ --api-key KEY

  # Classify a single symbol (test):
  python step4_sow_agent.py classify --memory output/sow_symbol_memory.json \\
      --symbol "FLOW TRANSMITTER" --description "circle with FT text"

  # Classify all tags from table extraction (Step 6):
  python step4_sow_agent.py filter \\
      --memory output/sow_symbol_memory.json \\
      --tags output/master_tags.json \\
      --out output/

  # Full pipeline (build + integrate with drawing context):
  python step4_sow_agent.py build --excel ANNEXURE-2.xlsx \\
      --context output/drawing_context.json --api-key KEY
"""

import argparse
import base64
import json
import logging
import os
import re
import sys
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

# ── Models ─────────────────────────────────────────────────────────────────────
GEMINI_FLASH_MODEL = "gemini-2.5-flash"
GEMINI_MAX_SIDE    = 512    # symbol images are small — no need for 4096

# ── ISA Category mapping (from description keywords) ──────────────────────────
_ISA_CATEGORY_MAP = [
    (["FLOW CONTROL", "FCV"],                   "Control Valve",       "FC"),
    (["TEMPERATURE CONTROL", "TCV"],            "Control Valve",       "TC"),
    (["PRESSURE CONTROL", "PCV"],               "Control Valve",       "PC"),
    (["SHUTDOWN VALVE", "ESDV", "SDV"],         "Safety/Shutdown",     "XV"),
    (["EMERGENCY SHUTDOWN"],                    "Safety/Shutdown",     "ESDV"),
    (["SOLENOID VALVE"],                        "Valve",               "SV"),
    (["MOTOR OPERATED"],                        "Valve",               "MOV"),
    (["BALL VALVE"],                            "Valve",               "BV"),
    (["BUTTERFLY VALVE"],                       "Valve",               "XV"),
    (["GATE VALVE"],                            "Valve",               "GV"),
    (["GLOBE VALVE"],                           "Valve",               "XV"),
    (["NEEDLE VALVE"],                          "Valve",               "NV"),
    (["PLUG VALVE"],                            "Valve",               "PV"),
    (["NON RETURN VALVE"],                      "Valve",               "NRV"),
    (["PRESSURE SAFETY VALVE", "PSV"],          "Safety/Shutdown",     "PSV"),
    (["SAFETY RELIEF VALVE", "SRV"],            "Safety/Shutdown",     "SRV"),
    (["ON OFF VALVE"],                          "Valve",               "XV"),
    (["CHOKE VALVE"],                           "Valve",               "CV"),
    (["QUICK EXHAUST"],                         "Valve",               "QEV"),
    (["FLOAT VALVE"],                           "Valve",               "FV"),
    (["WING VALVE"],                            "Valve",               "WV"),
    (["DOUBLE BLOCK"],                          "Valve",               "DBBV"),
    (["3-WAY SOLENOID"],                        "Valve",               "SV"),
    (["FLOW INDICATING TRANSMITTER", "FIT"],    "Transmitter",         "FIT"),
    (["FLOW TRANSMITTER"],                      "Transmitter",         "FT"),
    (["LEVEL INDICATING TRANSMITTER", "LIT"],   "Transmitter",         "LIT"),
    (["LEVEL TRANSMITTER"],                     "Transmitter",         "LT"),
    (["PRESSURE INDICATING TRANSMITTER","PIT"], "Transmitter",         "PIT"),
    (["PRESSURE TRANSMITTER"],                  "Transmitter",         "PT"),
    (["TEMPERATURE INDICATING TRANSMITTER"],    "Transmitter",         "TIT"),
    (["TEMPERATURE TRANSMITTER"],               "Transmitter",         "TT"),
    (["DIFFERENTIAL PRESSURE INDICATING"],      "Transmitter",         "DPIT"),
    (["CORROSION TRANSMITTER"],                 "Transmitter",         "CT"),
    (["DENSITY INDICATING TRANSMITTER"],        "Transmitter",         "DIT"),
    (["FLOW ELEMENT"],                          "Element",             "FE"),
    (["LEVEL ELEMENT"],                         "Element",             "LE"),
    (["TEMPERATURE ELEMENT"],                   "Element",             "TE"),
    (["ANALYZER ELEMENT"],                      "Element",             "AE"),
    (["CORROSION ELEMENT"],                     "Element",             "CE"),
    (["CORROSION COUPON"],                      "Corrosion Monitoring","CC"),
    (["CORROSION PROBE"],                       "Corrosion Monitoring","CP"),
    (["FLOW GAUGE"],                            "Gauge",               "FG"),
    (["LEVEL GAUGE"],                           "Gauge",               "LG"),
    (["PRESSURE GAUGE"],                        "Gauge",               "PG"),
    (["TEMPERATURE GAUGE"],                     "Gauge",               "TG"),
    (["DIFFERENTIAL PRESSURE GAUGE"],           "Gauge",               "DPG"),
    (["LEVEL INDICATOR"],                       "Indicator",           "LI"),
    (["PRESSURE INDICATOR"],                    "Indicator",           "PI"),
    (["TEMPERATURE INDICATOR"],                 "Indicator",           "TI"),
    (["LEVEL SWITCH"],                          "Switch",              "LS"),
    (["LIMIT SWITCH CLOSE"],                    "Switch",              "LSC"),
    (["LIMIT SWITCH OPEN"],                     "Switch",              "LSO"),
    (["LIMIT SWITCH"],                          "Switch",              "LS"),
    (["PRESSURE SWITCH"],                       "Switch",              "PS"),
    (["SELECTOR SWITCH"],                       "Switch",              "SS"),
    (["HAND SWITCH OPEN"],                      "Switch",              "HSO"),
    (["HAND SWITCH CLOSE"],                     "Switch",              "HSC"),
    (["FLOW TOTALISER", "FLOW TOTALIZER"],      "Totaliser",           "FQI"),
    (["ANALYZER"],                              "Analyzer",            "AT"),
    (["AIR COMPRESSOR"],                        "Compressor",          "KC"),
    (["CENTRIFUGAL PUMP"],                      "Pump",                "P"),
    (["RECIPROCATING PUMP"],                    "Pump",                "P"),
    (["ROTOR PUMP"],                            "Pump",                "P"),
    (["VERTICAL PUMP"],                         "Pump",                "P"),
    (["VERTICAL SUMP PUMP"],                    "Pump",                "P"),
    (["HEAT EXCHANGER"],                        "Heat Transfer",       "E"),
    (["AIR COOLER"],                            "Heat Transfer",       "E"),
    (["COALESCER"],                             "Separation",          "S"),
    (["FILTER STRAINER", "Y-TYPE FILTER"],      "Filter/Strainer",     "STR"),
    (["RESTRICTION ORIFICE"],                   "Flow Element",        "RO"),
    (["THERMOWELL"],                            "Temperature",         "TW"),
    (["SIGHT GLASS"],                           "Gauge",               "SG"),
    (["VESSEL"],                                "Vessel",              "V"),
    (["TANK"],                                  "Tank",                "TK"),
    (["FLOATING ROOF TANK"],                    "Tank",                "TK"),
    (["PIG LAUNCHER"],                          "Pipeline Pigging",    "PL"),
    (["PIG RECEIVER"],                          "Pipeline Pigging",    "PR"),
    (["PIG SIGNALER"],                          "Pipeline Pigging",    "PS"),
    (["MOTOR"],                                 "Electrical",          "M"),
    (["BLOWER"],                                "Rotating",            "BL"),
    (["STATIC MIXER"],                          "Mixing",              "MX"),
    (["FLAME ARRESTER"],                        "Safety",              "FA"),
    (["AIR DRYER"],                             "Utility",             "AD"),
    (["AIR RECEIVER"],                          "Utility",             "AR"),
    (["EXHAUST SILENCER"],                      "Utility",             "SIL"),
    (["INJECTION QUILL"],                       "Chemical Injection",  "QU"),
    (["I/P CONVERTER"],                         "Converter",           "IP"),
    (["ELECTRO-PNEUMATIC"],                     "Converter",           "EP"),
    (["INDICATION LAMP CLOSE"],                 "Indication",          "IL"),
    (["INDICATION LAMP OPEN"],                  "Indication",          "IL"),
    (["LIGHT ALARM"],                           "Alarm",               "LA"),
    (["BI-DIRECTIONAL POSITION"],               "Position",            "ZIT"),
    (["PIPELINE TAG"],                          "Pipeline",            "LINE"),
    (["INDUCTOR", "EJECTOR"],                   "Process",             "EJ"),
    (["FIRED HEATER"],                          "Heater",              "H"),
    (["CALIBRATION POT"],                       "Instrument Accessory","CAL"),
    (["COOLING FAN"],                           "Cooling",             "CF"),
    # DO NOT USE categories
    (["COMPUTER FUNCTION", "SIGNAL TAG"],       "Signal/Logic",        "FC"),
    (["DCS", "SIGNAL TAG"],                     "Signal/Logic",        ""),
    (["PLC", "SIGNAL TAG"],                     "Signal/Logic",        ""),
    (["NON MAINTAINABLE"],                      "Non-Maintainable",    ""),
    (["DESTINATION ARROW"],                     "Drawing Annotation",  ""),
    (["INSULATING COUPLING"],                   "Piping",              ""),
    (["FLANGE"],                                "Piping",              ""),
    (["TUNDISH"],                               "Piping",              ""),
    (["REDUCER"],                               "Piping",              ""),
    (["CAUSE AND EFFECTS"],                     "Logic/Control",       ""),
    (["SOFT TAG"],                              "Signal/Logic",        ""),
    (["INSULATION"],                            "Piping Annotation",   ""),
    (["BARRED TEE"],                            "Piping",              ""),
    (["MANWAY"],                                "Vessel Accessory",    ""),
    (["TIE-IN"],                                "Tie-In",              ""),
    (["DIAPHRAGM SEAL"],                        "Instrument Accessory",""),
    (["REVISION TRIANGLE"],                     "Drawing Annotation",  ""),
]

# ── ISA shape descriptors (visual features by category) ───────────────────────
_SHAPE_DESCRIPTORS = {
    "Transmitter":      ["circle", "bubble", "letter code inside circle", "line below"],
    "Gauge":            ["circle", "bubble", "letter code inside circle"],
    "Indicator":        ["circle", "bubble", "indicator letters"],
    "Switch":           ["circle", "bubble", "S prefix or letter"],
    "Element":          ["circle or small symbol", "element shape", "field instrument"],
    "Control Valve":    ["bow-tie shape", "triangle", "actuator on top", "valve body"],
    "Valve":            ["valve body shape", "line through symbol"],
    "Safety/Shutdown":  ["valve with actuator", "spring symbol", "safety marker"],
    "Pump":             ["circle with arrow", "rotating symbol"],
    "Compressor":       ["diamond or rectangle", "rotating machinery"],
    "Heat Transfer":    ["zigzag lines", "shell and tube symbol"],
    "Vessel":           ["tall rectangle", "vessel outline"],
    "Tank":             ["wide rectangle", "tank outline"],
    "Filter/Strainer":  ["Y-shape or mesh symbol", "filter body"],
    "Signal/Logic":     ["hexagon", "diamond", "dashed box", "logic symbol"],
    "Piping":           ["line annotation", "piping fitting symbol"],
    "Pipeline Pigging": ["arrow in pipe", "pig symbol"],
    "Corrosion Monitoring": ["box with CC or CP", "inline probe symbol"],
}


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _infer_category_and_prefix(desc: str) -> tuple[str, str]:
    """Return (isa_category, typical_tag_prefix) from description string."""
    desc_up = desc.upper()
    for keywords, category, prefix in _ISA_CATEGORY_MAP:
        if any(kw in desc_up for kw in keywords):
            return category, prefix
    return "Unknown", ""


def _infer_shape_features(desc: str, category: str) -> list[str]:
    """Return list of expected visual shape features."""
    base = _SHAPE_DESCRIPTORS.get(category, ["unknown shape"])
    extras = []
    desc_up = desc.upper()
    if "INDICATING" in desc_up:
        extras.append("indicating (I letter in tag)")
    if "RECORDING" in desc_up:
        extras.append("recording (R letter in tag)")
    if "SWITCH" in desc_up and "OPEN" in desc_up:
        extras.append("open position indicator")
    if "SWITCH" in desc_up and "CLOSE" in desc_up:
        extras.append("closed position indicator")
    if "3-WAY" in desc_up:
        extras.append("three-port valve body")
    return base + extras


def _safe_img_encode(img_bytes: bytes) -> Optional[str]:
    """Encode image bytes to base64 string."""
    if not img_bytes:
        return None
    try:
        return base64.b64encode(img_bytes).decode("utf-8")
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# Phase 1 — Parse Excel
# ═══════════════════════════════════════════════════════════════════════════════

def parse_excel_symbols(excel_path: str) -> list[dict]:
    """
    Read both sheets from the Excel file.
    Returns list of raw symbol dicts with embedded image bytes.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("pip install openpyxl")

    wb    = load_workbook(excel_path, read_only=False)
    sheet_cfg = {
        "SYMBOL-USE":        ("USE",        "ALLOW"),
        "SYMBOL DONT USE":   ("DO_NOT_USE", "BLOCK"),
    }

    symbols: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_cfg:
            log.warning("Unexpected sheet '%s' — skipping", sheet_name)
            continue

        sheet_src, status = sheet_cfg[sheet_name]
        ws = wb[sheet_name]

        # Build row → image bytes map
        row_to_imgs: dict[int, list[bytes]] = {}
        for img in (ws._images if hasattr(ws, "_images") else []):
            try:
                row_1idx = img.anchor._from.row + 1   # 0-indexed → 1-indexed
                img_data = img._data()
                if img_data:
                    row_to_imgs.setdefault(row_1idx, []).append(img_data)
            except Exception:
                pass

        # Iterate data rows (skip header row 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            sno  = row[0]
            desc = row[2] if len(row) > 2 else None
            if not desc or not str(desc).strip():
                continue

            sno_int  = int(sno) if sno is not None else 0
            row_1idx = sno_int + 1   # header at row 1, data at row sno+1

            # Images for this row (may be 1 or 2)
            img_list = row_to_imgs.get(row_1idx, [])

            category, prefix = _infer_category_and_prefix(str(desc))
            shape_features   = _infer_shape_features(str(desc), category)

            symbol = {
                "sno":                  sno_int,
                "symbol_name":          str(desc).strip(),
                "symbol_name_normalized": re.sub(r'\s+', ' ', str(desc).strip().upper()),
                "symbol_category":      category,
                "typical_tag_prefix":   prefix,
                "symbol_description":   f"{desc} — {category} instrument/equipment",
                "sheet_source":         sheet_src,
                "status":               status,
                "has_image":            len(img_list) > 0,
                "image_count":          len(img_list),
                "symbol_visual_features": shape_features,
                "image_b64":            [_safe_img_encode(b) for b in img_list],
                "gemini_visual_description": None,   # filled in Phase 2
            }
            symbols.append(symbol)

        log.info("Sheet '%s': %d symbols parsed (%s)", sheet_name, 
                 sum(1 for s in symbols if s["sheet_source"] == sheet_src), status)

    return symbols


# ═══════════════════════════════════════════════════════════════════════════════
# Phase 2 — Gemini vision enrichment
# ═══════════════════════════════════════════════════════════════════════════════

_SYMBOL_DESCRIBE_PROMPT = """You are an ISA 5.1 P&ID symbol expert.

Describe this engineering P&ID symbol image in precise technical terms.
Focus on:
  1. Shape (circle, hexagon, diamond, triangle, rectangle, bow-tie, etc.)
  2. Internal markings (letters, numbers, lines, hatching)
  3. External connections (lines entering/leaving, arrows)
  4. Actuator/accessories if visible
  5. Any distinctive visual identifier

The symbol is labeled: "{symbol_name}" (category: {category})

Return ONLY a JSON object (no markdown):
{{
  "primary_shape": "circle|hexagon|diamond|rectangle|bow-tie|valve-body|custom",
  "internal_markings": "e.g. 'FT' text, horizontal line dividing circle, dashes",
  "connection_lines": "e.g. 2 horizontal lines, 1 vertical, none",
  "distinctive_features": ["list", "of", "unique", "visual", "features"],
  "isa_match_hint": "one sentence describing how to visually identify this on a P&ID",
  "similar_to": "symbol it could be confused with, or none"
}}"""


def _build_gemini_client(api_key: str):
    try:
        import google.genai as genai
        return genai.Client(api_key=api_key), "new"
    except Exception:
        pass
    try:
        import google.generativeai as gl
        gl.configure(api_key=api_key)
        return gl, "legacy"
    except Exception as e:
        raise RuntimeError(f"No Gemini SDK: {e}")


def _gemini_image_call(client, sdk: str, model: str,
                        img_bytes: bytes, prompt: str) -> str:
    import cv2, numpy as np
    # Resize to max 512px (symbol images are tiny)
    img_arr = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), cv2.IMREAD_COLOR)
    if img_arr is not None:
        H, W = img_arr.shape[:2]
        if max(H, W) > GEMINI_MAX_SIDE:
            scale = GEMINI_MAX_SIDE / max(H, W)
            img_arr = cv2.resize(img_arr, (int(W*scale), int(H*scale)))
        _, buf = cv2.imencode(".png", img_arr)
        img_bytes = buf.tobytes()
        mime = "image/png"
    else:
        mime = "image/png"

    if sdk == "new":
        from google.genai import types as gt
        resp = client.models.generate_content(
            model=model,
            contents=[
                gt.Part.from_bytes(data=img_bytes, mime_type=mime),
                gt.Part.from_text(text=prompt),
            ],
        )
        return resp.text.strip()
    else:
        import google.generativeai as gl
        import PIL.Image as PILImage
        import io
        pil = PILImage.open(io.BytesIO(img_bytes))
        resp = gl.GenerativeModel(model).generate_content([prompt, pil])
        return resp.text.strip()


def _parse_json(raw: str) -> dict:
    clean = raw.replace("```json", "").replace("```", "").strip()
    m = re.search(r'\{.*\}', clean, re.DOTALL)
    return json.loads(m.group(0) if m else clean)


def enrich_with_gemini_vision(symbols: list[dict], client, sdk: str,
                               batch_size: int = 5) -> list[dict]:
    """
    Phase 2: For each symbol that has an image, call Gemini to describe it.
    Processes in batches to manage rate limits.
    """
    has_image = [s for s in symbols if s["has_image"] and s["image_b64"]]
    log.info("Enriching %d symbols with Gemini vision...", len(has_image))

    enriched = 0
    errors   = 0

    for i, sym in enumerate(has_image):
        img_b64 = sym["image_b64"][0]   # use first image
        if not img_b64:
            continue

        img_bytes = base64.b64decode(img_b64)
        prompt    = _SYMBOL_DESCRIBE_PROMPT.format(
            symbol_name=sym["symbol_name"],
            category=sym["symbol_category"],
        )

        try:
            raw    = _gemini_image_call(client, sdk, GEMINI_FLASH_MODEL,
                                        img_bytes, prompt)
            parsed = _parse_json(raw)
            sym["gemini_visual_description"] = parsed

            # Merge gemini features into visual_features
            gemini_features = parsed.get("distinctive_features", [])
            sym["symbol_visual_features"] = list(set(
                sym["symbol_visual_features"] + gemini_features
            ))
            enriched += 1

            if (i + 1) % 10 == 0:
                log.info("  Enriched %d/%d symbols...", i + 1, len(has_image))

        except json.JSONDecodeError as e:
            log.warning("JSON parse error for '%s': %s", sym["symbol_name"], e)
            sym["gemini_visual_description"] = {"raw_response": raw, "parse_error": str(e)}
            errors += 1
        except Exception as e:
            log.warning("Gemini failed for '%s': %s", sym["symbol_name"], e)
            sym["gemini_visual_description"] = {"error": str(e)}
            errors += 1

    log.info("Vision enrichment: %d OK, %d errors", enriched, errors)
    return symbols


# ═══════════════════════════════════════════════════════════════════════════════
# Phase 3 — Build Symbol Scope Memory
# ═══════════════════════════════════════════════════════════════════════════════

def build_symbol_scope_memory(symbols: list[dict]) -> dict:
    """
    Build the final Symbol Scope Memory object.
    This is the persistent knowledge base consumed by all downstream agents.
    """
    use_symbols    = [s for s in symbols if s["sheet_source"] == "USE"]
    block_symbols  = [s for s in symbols if s["sheet_source"] == "DO_NOT_USE"]

    # Build lookup indexes for fast matching
    name_to_symbol: dict[str, dict] = {}
    prefix_to_symbols: dict[str, list] = {}

    for sym in symbols:
        key = sym["symbol_name_normalized"]
        name_to_symbol[key] = sym
        prefix = sym.get("typical_tag_prefix", "")
        if prefix:
            prefix_to_symbols.setdefault(prefix, []).append(sym)

    # Category breakdown
    def _category_counts(sym_list):
        counts = {}
        for s in sym_list:
            cat = s["symbol_category"]
            counts[cat] = counts.get(cat, 0) + 1
        return dict(sorted(counts.items(), key=lambda x: -x[1]))

    memory = {
        "version":             "v1",
        "source_excel":        "",   # filled by caller
        "build_timestamp":     "",   # filled by caller
        "total_symbols":       len(symbols),
        "allow_count":         len(use_symbols),
        "block_count":         len(block_symbols),
        "category_breakdown": {
            "ALLOW": _category_counts(use_symbols),
            "BLOCK": _category_counts(block_symbols),
        },
        "symbols":             symbols,
        "name_index":          list(name_to_symbol.keys()),
        "blocked_names":       [s["symbol_name_normalized"] for s in block_symbols],
        "allowed_names":       [s["symbol_name_normalized"] for s in use_symbols],
        "prefix_index":        {k: [s["symbol_name"] for s in v]
                               for k, v in prefix_to_symbols.items()},
        "classification_rules": {
            "ALLOW":       "Symbol in USE sheet → extract=true, scope_status=IN_SCOPE",
            "BLOCK":       "Symbol in DO_NOT_USE sheet → extract=false, scope_status=OUT_OF_SCOPE",
            "UNSPECIFIED": "Symbol not found in either → extract=true, scope_status=UNSPECIFIED, flag=NO_SCOPE_DEFINITION_FOUND",
        },
        "downstream_usage": {
            "tag_detection":    "Filter detected symbols against blocked_names before OCR",
            "ocr_agent":        "Skip OCR for BLOCK symbols, mark as OUT_OF_SCOPE",
            "validation_agent": "Validate tag prefixes against prefix_index",
            "assembly_agent":   "Drop OUT_OF_SCOPE tags from final output",
        },
    }
    return memory


# ═══════════════════════════════════════════════════════════════════════════════
# Phase 4 — Runtime classifier
# ═══════════════════════════════════════════════════════════════════════════════

def classify_symbol(symbol_name: str, visual_description: str,
                    memory: dict) -> dict:
    """
    Classify a detected symbol against the scope memory.
    Returns a full classification result dict.

    Rules:
      A — Found in USE   → ALLOW  / IN_SCOPE
      B — Found in DONT  → BLOCK  / OUT_OF_SCOPE
      C — Not found      → ALLOW  / UNSPECIFIED + flag
    """
    query_norm = re.sub(r'\s+', ' ', symbol_name.strip().upper())

    # ── Exact name match ──────────────────────────────────────────────────────
    for sym in memory.get("symbols", []):
        if sym["symbol_name_normalized"] == query_norm:
            allowed = (sym["status"] == "ALLOW")
            return {
                "symbol_name":        symbol_name,
                "matched_name":       sym["symbol_name"],
                "match_type":         "exact",
                "status":             sym["status"],
                "sheet_source":       sym["sheet_source"],
                "symbol_category":    sym["symbol_category"],
                "typical_tag_prefix": sym.get("typical_tag_prefix", ""),
                "extract":            allowed,
                "scope_status":       "IN_SCOPE" if allowed else "OUT_OF_SCOPE",
                "sow_source":         sym["sheet_source"],
                "extraction_allowed": allowed,
                "reason":             f"Exact match in {sym['sheet_source']} sheet",
            }

    # ── Partial name match (substring) ────────────────────────────────────────
    best_match = None
    best_score = 0

    for sym in memory.get("symbols", []):
        sym_norm = sym["symbol_name_normalized"]
        # Check how many words of query match
        q_words  = set(query_norm.split())
        s_words  = set(sym_norm.split())
        overlap  = len(q_words & s_words)
        max_len  = max(len(q_words), len(s_words))
        score    = overlap / max_len if max_len > 0 else 0

        if score > best_score and score >= 0.6:   # 60% word overlap threshold
            best_score = score
            best_match = sym

    if best_match:
        allowed = (best_match["status"] == "ALLOW")
        return {
            "symbol_name":        symbol_name,
            "matched_name":       best_match["symbol_name"],
            "match_type":         f"partial (score={best_score:.2f})",
            "status":             best_match["status"],
            "sheet_source":       best_match["sheet_source"],
            "symbol_category":    best_match["symbol_category"],
            "typical_tag_prefix": best_match.get("typical_tag_prefix", ""),
            "extract":            allowed,
            "scope_status":       "IN_SCOPE" if allowed else "OUT_OF_SCOPE",
            "sow_source":         best_match["sheet_source"],
            "extraction_allowed": allowed,
            "reason":             f"Partial match ({best_score:.0%} word overlap) with '{best_match['symbol_name']}' in {best_match['sheet_source']}",
        }

    # ── Not found ─────────────────────────────────────────────────────────────
    return {
        "symbol_name":        symbol_name,
        "matched_name":       None,
        "match_type":         "not_found",
        "status":             "UNSPECIFIED",
        "sheet_source":       "NOT_FOUND",
        "symbol_category":    "Unknown",
        "typical_tag_prefix": "",
        "extract":            True,
        "scope_status":       "UNSPECIFIED",
        "sow_source":         "NOT_FOUND",
        "extraction_allowed": True,
        "reason":             "NO_SCOPE_DEFINITION_FOUND — extract with flag for human review",
    }


def check_tag_scope(tag_id: str, symbol_name: str,
                    memory: dict,
                    visual_description: str = "") -> dict:
    """
    Full tag scope check as required by Phase 4 output spec.
    Returns the standard output dict for every detected tag.
    """
    classification = classify_symbol(symbol_name, visual_description, memory)
    return {
        "tag_id":             tag_id,
        "symbol_name":        symbol_name,
        "symbol_type":        classification.get("symbol_category", "Unknown"),
        "scope_status":       classification["scope_status"],
        "sow_source":         classification["sow_source"],
        "extraction_allowed": classification["extraction_allowed"],
        "typical_tag_prefix": classification.get("typical_tag_prefix", ""),
        "match_type":         classification.get("match_type", "not_found"),
        "reason":             classification.get("reason", ""),
    }


def filter_tags_against_sow(master_tags: list[dict],
                             memory: dict) -> dict:
    """
    Filter a list of tag dicts (from step6 master_tags.json) against SOW memory.
    Returns: {allowed, blocked, unspecified, all_results}
    """
    allowed     : list[dict] = []
    blocked     : list[dict] = []
    unspecified : list[dict] = []

    for tag_entry in master_tags:
        tag_id      = tag_entry.get("tag_number", "")
        symbol_name = tag_entry.get("column_header", "") or tag_entry.get("symbol_name", "")
        table_type  = tag_entry.get("table_type", "")

        # If no symbol name, infer from tag prefix
        if not symbol_name or symbol_name == tag_id:
            m = re.match(r'^([A-Z]{1,5})', tag_id.upper())
            if m:
                prefix = m.group(1)
                candidates = memory.get("prefix_index", {}).get(prefix, [])
                symbol_name = candidates[0] if candidates else f"Tag prefix {prefix}"

        result = check_tag_scope(tag_id, symbol_name, memory)
        result["original_tag_entry"] = tag_entry

        scope = result["scope_status"]
        if scope == "OUT_OF_SCOPE":
            blocked.append(result)
        elif scope == "UNSPECIFIED":
            unspecified.append(result)
        else:
            allowed.append(result)

    log.info("SOW filter: %d allowed | %d blocked | %d unspecified",
             len(allowed), len(blocked), len(unspecified))

    return {
        "allowed":     allowed,
        "blocked":     blocked,
        "unspecified": unspecified,
        "summary": {
            "total":       len(master_tags),
            "allowed":     len(allowed),
            "blocked":     len(blocked),
            "unspecified": len(unspecified),
        },
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Scope summary builder
# ═══════════════════════════════════════════════════════════════════════════════

def build_scope_summary(memory: dict) -> str:
    """Human-readable scope summary for injection into downstream agent prompts."""
    lines = [
        "== SOW SYMBOL SCOPE MEMORY ==",
        f"Total symbols: {memory['total_symbols']} "
        f"({memory['allow_count']} ALLOW | {memory['block_count']} BLOCK)",
        "",
        "-- ALLOWED SYMBOLS (extract these) --",
    ]
    for sym in memory.get("symbols", []):
        if sym["status"] == "ALLOW":
            prefix = f"[{sym['typical_tag_prefix']}]" if sym.get("typical_tag_prefix") else ""
            lines.append(f"  ✓ {prefix:>8}  {sym['symbol_name']}")

    lines += ["", "-- BLOCKED SYMBOLS (do NOT extract) --"]
    for sym in memory.get("symbols", []):
        if sym["status"] == "BLOCK":
            lines.append(f"  ✗           {sym['symbol_name']}")

    lines += [
        "",
        "-- CLASSIFICATION RULES --",
        "  IF symbol in ALLOW list  → extract=true,  scope_status=IN_SCOPE",
        "  IF symbol in BLOCK list  → extract=false, scope_status=OUT_OF_SCOPE",
        "  IF symbol not found      → extract=true,  scope_status=UNSPECIFIED (flag for review)",
        "",
        "== END SOW MEMORY ==",
    ]
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# Main pipeline
# ═══════════════════════════════════════════════════════════════════════════════

def run_build(excel_path: str, out_dir: str, api_key: Optional[str],
              drawing_context_path: Optional[str] = None,
              skip_vision: bool = False) -> dict:
    """Build the Symbol Scope Memory from Excel."""
    import datetime
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    log.info("=== Phase 1: Parsing Excel ===")
    symbols = parse_excel_symbols(excel_path)
    log.info("Parsed %d total symbols (%d ALLOW + %d BLOCK)",
             len(symbols),
             sum(1 for s in symbols if s["status"] == "ALLOW"),
             sum(1 for s in symbols if s["status"] == "BLOCK"))

    # Phase 2: Vision enrichment (optional, requires API key)
    if api_key and not skip_vision:
        log.info("=== Phase 2: Gemini Vision Enrichment ===")
        client, sdk = _build_gemini_client(api_key)
        symbols = enrich_with_gemini_vision(symbols, client, sdk)
    else:
        log.info("Phase 2 skipped (no API key or --skip-vision)")

    log.info("=== Phase 3: Building Symbol Scope Memory ===")
    memory = build_symbol_scope_memory(symbols)
    memory["source_excel"]      = str(Path(excel_path).resolve())
    memory["build_timestamp"]   = datetime.datetime.now().isoformat()

    # Remove image b64 blobs from memory JSON (too large; keep in symbols but strip for export)
    memory_export = json.loads(json.dumps(memory))
    for sym in memory_export.get("symbols", []):
        sym.pop("image_b64", None)   # strip blobs from JSON export

    # Write sow_symbol_memory.json
    memory_path = str(out / "sow_symbol_memory.json")
    with open(memory_path, "w") as f:
        json.dump(memory_export, f, indent=2)
    log.info("✓ sow_symbol_memory.json (%d symbols) → %s",
             len(symbols), memory_path)

    # Write scope_summary.txt
    summary_text = build_scope_summary(memory_export)
    summary_path = str(out / "sow_scope_summary.txt")
    with open(summary_path, "w") as f:
        f.write(summary_text)
    log.info("✓ sow_scope_summary.txt → %s", summary_path)

    # Update drawing_context.json
    ctx_path = drawing_context_path or str(out / "drawing_context.json")
    if Path(ctx_path).exists():
        with open(ctx_path) as f:
            dctx = json.load(f)
        dctx["sow_memory_path"]   = memory_path
        dctx["sow_summary_path"]  = summary_path
        dctx["sow_summary"] = {
            "allow_count":   memory["allow_count"],
            "block_count":   memory["block_count"],
            "total_symbols": memory["total_symbols"],
            "categories_allowed": list(memory["category_breakdown"]["ALLOW"].keys()),
        }
        with open(ctx_path, "w") as f:
            json.dump(dctx, f, indent=2)
        log.info("✓ drawing_context.json updated")

    return memory_export


def run_filter(memory_path: str, tags_path: str, out_dir: str) -> dict:
    """Filter master_tags.json against the SOW memory."""
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    with open(memory_path) as f:
        memory = json.load(f)
    with open(tags_path) as f:
        master_tags = json.load(f)

    log.info("Filtering %d tags against SOW memory (%d symbols)...",
             len(master_tags), memory.get("total_symbols", 0))

    results  = filter_tags_against_sow(master_tags, memory)

    # Write filtered results
    filtered_path = str(out / "sow_filtered_tags.json")
    with open(filtered_path, "w") as f:
        json.dump(results, f, indent=2)
    log.info("✓ sow_filtered_tags.json → %s", filtered_path)

    # Write final extractable tags only
    extractable = results["allowed"] + results["unspecified"]
    final_path  = str(out / "final_extractable_tags.json")
    with open(final_path, "w") as f:
        json.dump([r for r in extractable], f, indent=2)
    log.info("✓ final_extractable_tags.json (%d tags) → %s",
             len(extractable), final_path)

    return results


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Step 4: SOW Symbol Scope Intelligence Agent")
    sub = parser.add_subparsers(dest="cmd", required=True)

    # --- build ---
    build_p = sub.add_parser("build",
        help="Parse Excel + build Symbol Scope Memory")
    build_p.add_argument("--excel",   required=True,
                         help="Path to ANNEXURE-2_CDC-SYMBOLS_USED_AND_NOT_USED.xlsx")
    build_p.add_argument("--out",     default="output")
    build_p.add_argument("--api-key", help="Gemini API key for vision enrichment")
    build_p.add_argument("--context", help="drawing_context.json to update")
    build_p.add_argument("--skip-vision", action="store_true",
                         help="Skip Gemini vision enrichment (faster, text-only)")

    # --- classify ---
    cls_p = sub.add_parser("classify",
        help="Classify a single symbol name")
    cls_p.add_argument("--memory",      required=True, help="sow_symbol_memory.json")
    cls_p.add_argument("--symbol",      required=True, help="Symbol name to classify")
    cls_p.add_argument("--description", default="",    help="Visual description (optional)")

    # --- filter ---
    flt_p = sub.add_parser("filter",
        help="Filter master_tags.json from Step 6 against SOW")
    flt_p.add_argument("--memory", required=True, help="sow_symbol_memory.json")
    flt_p.add_argument("--tags",   required=True, help="master_tags.json from Step 6")
    flt_p.add_argument("--out",    default="output")

    args = parser.parse_args()

    if args.cmd == "build":
        api_key = (args.api_key
                   or os.environ.get("GEMINI_API_KEY")
                   or os.environ.get("GOOGLE_API_KEY"))
        if not api_key and not args.skip_vision:
            log.info("No API key — running in text-only mode (use --api-key for vision enrichment)")

        memory = run_build(
            excel_path           = args.excel,
            out_dir              = args.out,
            api_key              = api_key,
            drawing_context_path = getattr(args, "context", None),
            skip_vision          = args.skip_vision or not api_key,
        )

        print("\n=== Step 4 Complete — SOW Symbol Scope Memory Built ===")
        print(f"  Total symbols      : {memory['total_symbols']}")
        print(f"  ALLOW (USE)        : {memory['allow_count']}")
        print(f"  BLOCK (DO NOT USE) : {memory['block_count']}")
        print(f"\n  ALLOW categories:")
        for cat, cnt in list(memory["category_breakdown"]["ALLOW"].items())[:10]:
            print(f"    {cat:<32} {cnt:>3}")
        print(f"\n  BLOCK categories:")
        for cat, cnt in list(memory["category_breakdown"]["BLOCK"].items()):
            print(f"    {cat:<32} {cnt:>3}")
        print(f"\n  Output:")
        print(f"    {args.out}/sow_symbol_memory.json")
        print(f"    {args.out}/sow_scope_summary.txt")

    elif args.cmd == "classify":
        with open(args.memory) as f:
            memory = json.load(f)
        result = check_tag_scope("TEST-001", args.symbol, memory, args.description)
        print(json.dumps(result, indent=2))

    elif args.cmd == "filter":
        results = run_filter(args.memory, args.tags, args.out)
        s = results["summary"]
        print(f"\n=== SOW Filter Complete ===")
        print(f"  Total tags    : {s['total']}")
        print(f"  Allowed       : {s['allowed']}")
        print(f"  Blocked       : {s['blocked']}")
        print(f"  Unspecified   : {s['unspecified']}")
        print(f"\n  Output: {args.out}/sow_filtered_tags.json")
        print(f"          {args.out}/final_extractable_tags.json")


if __name__ == "__main__":
    main()
